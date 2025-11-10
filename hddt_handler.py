import io
import re
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook, Workbook

# --- Các hàm tiện ích nội bộ ---
def _clean_string_hddt(s):
    """Làm sạch chuỗi, loại bỏ khoảng trắng thừa và ký tự '."""
    if s is None: return ""
    cleaned_s = str(s).strip()
    if cleaned_s.startswith("'"): cleaned_s = cleaned_s[1:]
    return re.sub(r'\s+', ' ', cleaned_s)

def _to_float_hddt(value):
    """Chuyển đổi giá trị sang float, xử lý các trường hợp lỗi."""
    if value is None: return 0.0
    try:
        return float(str(value).replace(',', '').strip())
    except (ValueError, TypeError): return 0.0

def _format_tax_code_hddt(raw_vat_value):
    """Định dạng mã thuế."""
    if raw_vat_value is None: return ""
    try:
        s_value = str(raw_vat_value).replace('%', '').strip()
        f_value = float(s_value)
        if 0 < f_value < 1: f_value *= 100
        return f"{round(f_value):02d}"
    except (ValueError, TypeError): return ""
    
def _create_upsse_workbook_hddt():
    """Tạo một workbook Excel mới với các header chuẩn cho UpSSE."""
    headers = ["Mã khách", "Tên khách hàng", "Ngày", "Số hóa đơn", "Ký hiệu", "Diễn giải", "Mã hàng", "Tên mặt hàng", "Đvt", "Mã kho", "Mã vị trí", "Mã lô", "Số lượng", "Giá bán", "Tiền hàng", "Mã nt", "Tỷ giá", "Mã thuế", "Tk nợ", "Tk doanh thu", "Tk giá vốn", "Tk thuế có", "Cục thuế", "Vụ việc", "Bộ phận", "Lsx", "Sản phẩm", "Hợp đồng", "Phí", "Khế ước", "Nhân viên bán", "Tên KH(thuế)", "Địa chỉ (thuế)", "Mã số Thuế", "Nhóm Hàng", "Ghi chú", "Tiền thuế"]
    wb = Workbook()
    ws = wb.active
    for _ in range(4): ws.append([''] * len(headers))
    ws.append(headers)
    return wb

# --- Hàm tạo dòng BVMT (chỉ dùng cho hóa đơn riêng lẻ) ---
def _create_hddt_bvmt_row(original_row, phi_bvmt, static_data_hddt, khu_vuc):
    """Tạo dòng Thuế Bảo vệ Môi trường (BVMT) cho hóa đơn riêng lẻ."""
    bvmt_row = list(original_row)
    so_luong = _to_float_hddt(original_row[12])
    thue_suat = _to_float_hddt(original_row[17]) / 100.0 if original_row[17] else 0.0
    
    tien_hang_dong_bvmt = round(phi_bvmt * so_luong)
    tien_thue_dong_bvmt = round(tien_hang_dong_bvmt * thue_suat)

    bvmt_row[6], bvmt_row[7] = "TMT", "Thuế bảo vệ môi trường"
    bvmt_row[13] = phi_bvmt
    bvmt_row[14] = tien_hang_dong_bvmt
    bvmt_row[36] = tien_thue_dong_bvmt
    
    bvmt_row[18] = static_data_hddt.get('tk_no_bvmt_map', {}).get(khu_vuc)
    bvmt_row[19] = static_data_hddt.get('tk_dt_thue_bvmt_map', {}).get(khu_vuc)
    bvmt_row[20] = static_data_hddt.get('tk_gia_von_bvmt_value')
    bvmt_row[21] = static_data_hddt.get('tk_thue_co_bvmt_map', {}).get(khu_vuc)
    
    for i in [5, 31, 32, 33]: bvmt_row[i] = ''
    return bvmt_row

def _parse_date_from_excel_cell(date_val):
    """
    Attempts to parse a date value from an Excel cell, handling various formats.
    Returns a datetime.date object if successful, None otherwise.
    """
    if date_val is None:
        return None

    if isinstance(date_val, datetime):
        return date_val.date()

    if isinstance(date_val, (int, float)):
        try:
            return pd.to_datetime(float(date_val), unit='D', origin='1899-12-30').date()
        except (ValueError, TypeError):
            pass

    if isinstance(date_val, str):
        date_str = date_val.strip()
        date_formats = [
            '%d/%m/%Y', # 13/07/2025
            '%d-%m-%Y', # 13-07-2025
            '%Y/%m/%d', # 2025/07/13
            '%Y-%m-%d', # 2025-07-13
            '%d/%m/%y', # 13/07/25
            '%d-%m-%y', # 13-07-25
        ]
        for fmt in date_formats:
            try:
                return datetime.strptime(date_str, fmt).date()
            except ValueError:
                continue
    
    return None

# --- Hàm xử lý chính ---
def _generate_upsse_from_hddt_rows(rows_to_process, static_data_hddt, selected_chxd, final_date, summary_suffix_map):
    """Tạo các dòng dữ liệu cho file UpSSE từ dữ liệu bảng kê HĐĐT."""
    upsse_wb = _create_upsse_workbook_hddt()
    ws = upsse_wb.active

    if not rows_to_process:
        print(f"DEBUG: Không có dòng nào để xử lý trong giai đoạn này. Trả về workbook rỗng.")
        output_buffer = io.BytesIO()
        upsse_wb.save(output_buffer)
        output_buffer.seek(0)
        return output_buffer

    khu_vuc, ma_kho = static_data_hddt['chxd_to_khuvuc_map'].get(selected_chxd), static_data_hddt['tk_mk'].get(selected_chxd)
    # THÊM: lấy "Mã khách CHXD" từ cấu hình
    ma_khach_chxd = static_data_hddt.get('chxd_makh_map', {}).get(selected_chxd)
    tk_no, tk_doanh_thu, tk_gia_von, tk_thue_co = static_data_hddt['tk_no_map'].get(khu_vuc), static_data_hddt['tk_doanh_thu_map'].get(khu_vuc), static_data_hddt['tk_gia_von_value'], static_data_hddt['tk_thue_co_map'].get(khu_vuc)
    original_invoice_rows, bvmt_rows, summary_data = [], [], {}
    first_invoice_prefix_source = ""
    
    processed_row_count = 0
    for bkhd_row in rows_to_process:
        if _to_float_hddt(bkhd_row[9] if len(bkhd_row) > 9 else None) <= 0: 
            continue
        
        processed_row_count += 1
        ten_kh, ten_mat_hang = _clean_string_hddt(bkhd_row[4]), _clean_string_hddt(bkhd_row[7])
        is_anonymous, is_petrol = ("không lấy hóa đơn" in ten_kh.lower()), (ten_mat_hang in static_data_hddt['phi_bvmt_map'])
        
        if not is_anonymous or not is_petrol:
            new_upsse_row = [''] * 37
            new_upsse_row[9], new_upsse_row[1], new_upsse_row[31], new_upsse_row[2] = ma_kho, ten_kh, ten_kh, final_date
            so_hd_goc = str(bkhd_row[20] or '').strip()

            # --- LOGIC MỚI: SỐ HÓA ĐƠN CHO HÓA ĐƠN ĐỊNH DANH ---
            ky_hieu_str = str(bkhd_row[19] or '').strip()
            yy = ky_hieu_str[1:3] if len(ky_hieu_str) >= 3 else ''
            old_prefix = 'HN' if selected_chxd == 'Nguyễn Huệ' else ky_hieu_str[-2:]
            new_upsse_row[3] = f"{yy}{old_prefix}{so_hd_goc[-6:]}"

            new_upsse_row[4] = _clean_string_hddt(bkhd_row[18]) + _clean_string_hddt(bkhd_row[19])
            new_upsse_row[5], new_upsse_row[7], new_upsse_row[6] = f"Xuất bán hàng theo hóa đơn số {new_upsse_row[3]}", ten_mat_hang, static_data_hddt['ma_hang_map'].get(ten_mat_hang, '')
            new_upsse_row[8], new_upsse_row[12] = _clean_string_hddt(bkhd_row[11]), round(_to_float_hddt(bkhd_row[9]), 3)
            phi_bvmt = static_data_hddt['phi_bvmt_map'].get(ten_mat_hang, 0.0) if is_petrol else 0.0
            new_upsse_row[13] = _to_float_hddt(bkhd_row[10]) - phi_bvmt
            ma_thue = _format_tax_code_hddt(bkhd_row[15])
            new_upsse_row[17] = ma_thue
            thue_suat = _to_float_hddt(ma_thue) / 100.0 if ma_thue else 0.0
            tien_thue_goc, so_luong = _to_float_hddt(bkhd_row[16]), _to_float_hddt(bkhd_row[9])
            tien_thue_phi_bvmt = round(phi_bvmt * so_luong * thue_suat)
            new_upsse_row[36] = round(tien_thue_goc - tien_thue_phi_bvmt)
            new_upsse_row[14] = round(_to_float_hddt(bkhd_row[14]) if not is_petrol else _to_float_hddt(bkhd_row[17]) - tien_thue_goc - round(phi_bvmt * so_luong))
            new_upsse_row[18], new_upsse_row[19], new_upsse_row[20], new_upsse_row[21] = tk_no, tk_doanh_thu, tk_gia_von, tk_thue_co
            chxd_vu_viec_map = static_data_hddt['vu_viec_map'].get(selected_chxd, {})
            new_upsse_row[23] = chxd_vu_viec_map.get(ten_mat_hang, chxd_vu_viec_map.get("Dầu mỡ nhờn", ''))
            new_upsse_row[32], mst_khach_hang = _clean_string_hddt(bkhd_row[5]), _clean_string_hddt(bkhd_row[6])
            new_upsse_row[33] = mst_khach_hang
            ma_kh_fast = _clean_string_hddt(bkhd_row[2])
            # ĐỔI FALLBACK CUỐI: dùng "Mã khách CHXD" thay vì "ma_kho"
            new_upsse_row[0] = ma_kh_fast if ma_kh_fast and len(ma_kh_fast) < 12 else static_data_hddt['mst_to_makh_map'].get(mst_khach_hang, ma_khach_chxd)
            original_invoice_rows.append(new_upsse_row)
            if is_petrol: bvmt_rows.append(_create_hddt_bvmt_row(new_upsse_row, phi_bvmt, static_data_hddt, khu_vuc))
        
        else:
            if not first_invoice_prefix_source: first_invoice_prefix_source = str(bkhd_row[19] or '').strip()
            if ten_mat_hang not in summary_data:
                summary_data[ten_mat_hang] = {'sl': 0, 'thue': 0, 'phai_thu': 0, 'first_data': {'mau_so': _clean_string_hddt(bkhd_row[18]),'ky_hieu': _clean_string_hddt(bkhd_row[19]),'don_gia': _to_float_hddt(bkhd_row[10]),'vat_raw': bkhd_row[15]}}
            summary_data[ten_mat_hang]['sl'] += _to_float_hddt(bkhd_row[9])
            summary_data[ten_mat_hang]['thue'] += _to_float_hddt(bkhd_row[16])
            summary_data[ten_mat_hang]['phai_thu'] += _to_float_hddt(bkhd_row[17])
    
    prefix = first_invoice_prefix_source[-2:] if len(first_invoice_prefix_source) >= 2 else first_invoice_prefix_source
    for product, data in summary_data.items():
        summary_row = [''] * 37
        first_data = data['first_data']
        
        total_phai_thu = data['phai_thu']
        total_tien_thue_gtgt = data['thue']
        total_so_luong = data['sl']
        phi_bvmt_unit = static_data_hddt['phi_bvmt_map'].get(product, 0.0)
        ma_thue_str = _format_tax_code_hddt(first_data['vat_raw'])
        thue_suat = _to_float_hddt(ma_thue_str) / 100.0 if ma_thue_str else 0.0

        tien_hang_dong_bvmt = round(phi_bvmt_unit * total_so_luong)
        tien_thue_dong_bvmt = round(tien_hang_dong_bvmt * thue_suat)
        tien_thue_dong_goc = total_tien_thue_gtgt - tien_thue_dong_bvmt
        tien_hang_dong_goc = total_phai_thu - tien_hang_dong_bvmt - tien_thue_dong_bvmt - tien_thue_dong_goc
        
        # ĐỔI: cột A cho vãng lai dùng "Mã khách CHXD"
        summary_row[0], summary_row[1] = (ma_khach_chxd or ''), f"Khách hàng mua {product} không lấy hóa đơn"
        summary_row[31], summary_row[2] = summary_row[1], final_date

        # --- LOGIC MỚI: SỐ HÓA ĐƠN CHO HÓA ĐƠN TỔNG (VÃNG LAI) ---
        ky_hieu_any = str(first_data.get('ky_hieu', '') or '').strip()
        yy_vl = ky_hieu_any[1:3] if len(ky_hieu_any) >= 3 else ''
        summary_row[3] = f"{prefix}{yy_vl}.{final_date.strftime('%d.%m')}.{summary_suffix_map.get(product, '')}"

        summary_row[4] = first_data['mau_so'] + first_data['ky_hieu']
        summary_row[5] = f"Xuất bán hàng theo hóa đơn số {summary_row[3]}"
        summary_row[7], summary_row[6], summary_row[8], summary_row[9] = product, static_data_hddt['ma_hang_map'].get(product, ''), "Lít", ma_kho
        summary_row[12] = round(total_so_luong, 3)
        summary_row[13] = first_data['don_gia'] - phi_bvmt_unit
        summary_row[17] = ma_thue_str
        summary_row[14] = tien_hang_dong_goc
        summary_row[36] = tien_thue_dong_goc
        summary_row[18], summary_row[19], summary_row[20], summary_row[21] = tk_no, tk_doanh_thu, tk_gia_von, tk_thue_co
        summary_row[23] = static_data_hddt['vu_viec_map'].get(selected_chxd, {}).get(product, '')
        original_invoice_rows.append(summary_row)
        
        bvmt_summary_row = list(summary_row)
        bvmt_summary_row[6], bvmt_summary_row[7] = "TMT", "Thuế bảo vệ môi trường"
        bvmt_summary_row[13] = phi_bvmt_unit
        bvmt_summary_row[18] = static_data_hddt.get('tk_no_bvmt_map', {}).get(khu_vuc)
        bvmt_summary_row[19] = static_data_hddt.get('tk_dt_thue_bvmt_map', {}).get(khu_vuc)
        bvmt_summary_row[20] = static_data_hddt.get('tk_gia_von_bvmt_value')
        bvmt_summary_row[21] = static_data_hddt.get('tk_thue_co_bvmt_map', {}).get(khu_vuc)
        bvmt_summary_row[14] = tien_hang_dong_bvmt
        bvmt_summary_row[36] = tien_thue_dong_bvmt
        for i in [5, 31, 32, 33]: bvmt_summary_row[i] = ''
        bvmt_rows.append(bvmt_summary_row)
    
    for row_data in original_invoice_rows + bvmt_rows:
        ws.append(row_data)

    print(f"DEBUG: Số dòng hóa đơn gốc được thêm vào workbook: {len(original_invoice_rows)}")
    print(f"DEBUG: Số dòng BVMT được thêm vào workbook: {len(bvmt_rows)}")
    print(f"DEBUG: Tổng số dòng (sau khi lọc số lượng <= 0) được xử lý: {processed_row_count}")

    for row_index in range(6, ws.max_row + 1):
        date_cell = ws[f'C{row_index}']
        if isinstance(date_cell.value, datetime):
            date_cell.number_format = 'dd/mm/yyyy'
        text_cell = ws[f'R{row_index}']
        text_cell.number_format = '@'
        
    output_buffer = io.BytesIO()
    upsse_wb.save(output_buffer)
    output_buffer.seek(0)
    return output_buffer

# --- Khối lệnh điều phối chính ---
def process_hddt_report(file_content_bytes, selected_chxd, price_periods, new_price_invoice_number, confirmed_date_str=None, static_data_hddt=None, selected_chxd_symbol=None):
    """
    Xử lý bảng kê HĐĐT để tạo file UpSSE.
    """
    if static_data_hddt is None:
        raise ValueError("Dữ liệu cấu hình tĩnh cho HDDT chưa được tải.")
    if selected_chxd_symbol is None:
        raise ValueError("Ký hiệu hóa đơn của CHXD chưa được cung cấp để xác thực.")

    bkhd_wb = load_workbook(io.BytesIO(file_content_bytes), data_only=True)
    bkhd_ws = bkhd_wb.active

    if len(selected_chxd_symbol) < 6:
        raise ValueError(f"Ký hiệu hóa đơn trong file cấu hình ('{selected_chxd_symbol}') quá ngắn.")
    expected_invoice_symbol_suffix = selected_chxd_symbol[-6:].upper()

    has_at_least_one_valid_invoice_for_symbol_check = False
    
    max_rows_to_check = min(bkhd_ws.max_row, 100)
    for row_index, row_values in enumerate(bkhd_ws.iter_rows(min_row=11, max_row=max_rows_to_check, values_only=True), start=11):
        quantity_val = _to_float_hddt(row_values[9] if len(row_values) > 9 else None)
        
        if quantity_val <= 0:
            continue

        has_at_least_one_valid_invoice_for_symbol_check = True

        if len(row_values) > 19 and row_values[19] is not None:
            actual_invoice_symbol_hddt = _clean_string_hddt(row_values[19])
            if len(actual_invoice_symbol_hddt) >= 6:
                if actual_invoice_symbol_hddt[-6:].upper() != expected_invoice_symbol_suffix:
                    raise ValueError("Bảng kê HĐĐT không phải của cửa hàng bạn chọn.")
            else:
                raise ValueError(f"Ký hiệu hóa đơn tại dòng {row_index} của bảng kê quá ngắn.")
        else:
            raise ValueError(f"Hóa đơn tại dòng {row_index} của bảng kê thiếu ký hiệu hóa đơn (cột S).")
    
    if not has_at_least_one_valid_invoice_for_symbol_check:
        raise ValueError("Không tìm thấy hóa đơn hợp lệ nào trong file Bảng kê HDDT để xác thực.")

    final_date = None
    if confirmed_date_str:
        final_date = datetime.strptime(confirmed_date_str, '%Y-%m-%d')
    else:
        unique_dates = set()
        for row in bkhd_ws.iter_rows(min_row=11, values_only=True):
            quantity_val = _to_float_hddt(row[9] if len(row) > 9 else None)
            if quantity_val > 0:
                date_val_from_cell = row[21] if len(row) > 21 else None
                parsed_date = _parse_date_from_excel_cell(date_val_from_cell)
                if parsed_date:
                    unique_dates.add(parsed_date)
                else:
                    print(f"WARNING: Could not parse date '{date_val_from_cell}' from valid row.")
        
        if not unique_dates:
            raise ValueError("Không tìm thấy dữ liệu hóa đơn hợp lệ nào trong file Bảng kê HDDT.")
        
        if len(unique_dates) > 1:
            raise ValueError("Công cụ chỉ chạy được khi bạn kết xuất hóa đơn trong 1 ngày duy nhất.")
        
        the_date = unique_dates.pop()
        
        if the_date.day > 12:
            final_date = datetime(the_date.year, the_date.month, the_date.day)
        else:
            date1 = datetime(the_date.year, the_date.month, the_date.day)
            date2 = datetime(the_date.year, the_date.day, the_date.month)
            
            if date1 != date2:
                options = [
                    {'text': date1.strftime('%d/%m/%Y'), 'value': date1.strftime('%Y-%m-%d')},
                    {'text': date2.strftime('%d/%m/%Y'), 'value': date2.strftime('%Y-%m-%d')}
                ]
                options.sort(key=lambda x: datetime.strptime(x['value'], '%Y-%m-%d'))
                return {'choice_needed': True, 'options': options}
            else:
                final_date = date1

    all_rows = list(bkhd_ws.iter_rows(min_row=11, values_only=True))
    print(f"DEBUG: Tổng số dòng đọc được từ file Excel: {len(all_rows)}")

    # --- START: LOGIC CẬP NHẬT ---
    # Lấy danh sách mặt hàng xăng dầu từ dữ liệu cấu hình
    petroleum_products = static_data_hddt.get("petroleum_products", [])
    if not petroleum_products:
        print("WARNING: Không tìm thấy mặt hàng nào được đánh dấu là 'Xăng dầu' trong file MaHH.xlsx.")

    # Tạo suffix map một cách linh động
    # Hậu tố cho giai đoạn giá cũ sẽ bắt đầu từ 1
    suffix_map_old = {product: str(i + 1) for i, product in enumerate(petroleum_products)}
    
    # Hậu tố cho giai đoạn giá mới sẽ bắt đầu từ số lượng mặt hàng + 1
    # Ví dụ: có 4 mặt hàng, giá mới sẽ bắt đầu từ 5.
    # Thêm một khoảng đệm nhỏ để dễ phân biệt, ví dụ làm tròn lên 5.
    new_price_start_index = len(petroleum_products) + 1
    if new_price_start_index < 5: new_price_start_index = 5 # Đảm bảo bắt đầu ít nhất từ 5
    suffix_map_new = {product: str(i + new_price_start_index) for i, product in enumerate(petroleum_products)}
    # --- END: LOGIC CẬP NHẬT ---

    if price_periods == '1':
        print(f"DEBUG: Xử lý 1 giai đoạn giá. Suffix map: {suffix_map_old}")
        return _generate_upsse_from_hddt_rows(all_rows, static_data_hddt, selected_chxd, final_date, suffix_map_old)
    else:
        print(f"DEBUG: Xử lý 2 giai đoạn giá.")
        if not new_price_invoice_number: raise ValueError("Vui lòng nhập 'Số hóa đơn đầu tiên của giá mới'.")
        split_index = -1
        for i, row in enumerate(all_rows):
            if str(row[20] or '').strip() == new_price_invoice_number:
                split_index = i
                break
        
        print(f"DEBUG: Số hóa đơn giá mới cần tìm: '{new_price_invoice_number}'")
        print(f"DEBUG: split_index tìm được: {split_index}")

        if split_index == -1: 
            raise ValueError(f"Không tìm thấy hóa đơn số '{new_price_invoice_number}'.")
        
        rows_old_price = all_rows[:split_index]
        rows_new_price = all_rows[split_index:]

        print(f"DEBUG: Số dòng giá cũ: {len(rows_old_price)}. Suffix map: {suffix_map_old}")
        print(f"DEBUG: Số dòng giá mới: {len(rows_new_price)}. Suffix map: {suffix_map_new}")
        
        result_old = _generate_upsse_from_hddt_rows(rows_old_price, static_data_hddt, selected_chxd, final_date, suffix_map_old)
        result_new = _generate_upsse_from_hddt_rows(rows_new_price, static_data_hddt, selected_chxd, final_date, suffix_map_new)
        
        output_dict = {}
        if result_old: 
            result_old.seek(0)
            output_dict['old'] = result_old
        if result_new: 
            result_new.seek(0)
            output_dict['new'] = result_new
        
        if 'old' not in output_dict:
            empty_wb_old = _create_upsse_workbook_hddt()
            empty_buffer_old = io.BytesIO()
            empty_wb_old.save(empty_buffer_old)
            empty_buffer_old.seek(0)
            output_dict['old'] = empty_buffer_old
        
        if 'new' not in output_dict:
            empty_wb_new = _create_upsse_workbook_hddt()
            empty_buffer_new = io.BytesIO()
            empty_wb_new.save(empty_buffer_new)
            empty_buffer_new.seek(0)
            output_dict['new'] = empty_buffer_new

        return output_dict
