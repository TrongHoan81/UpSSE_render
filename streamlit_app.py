import streamlit as st
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import NamedStyle, Alignment
from datetime import datetime
import io
import os
import re # Import regex module

# --- Cấu hình trang Streamlit ---
st.set_page_config(layout="centered", page_title="Đồng bộ dữ liệu SSE")

# Đường dẫn đến các file cần thiết (giả định cùng thư mục với script)
LOGO_PATH = "Logo.png"
DATA_FILE_PATH = "Data.xlsx" # Tên chính xác của file dữ liệu

# Định nghĩa tiêu đề cho file UpSSE.xlsx (Di chuyển lên đây để luôn có sẵn)
headers = ["Mã khách", "Tên khách hàng", "Ngày", "Số hóa đơn", "Ký hiệu", "Diễn giải", "Mã hàng", "Tên mặt hàng",
           "Đvt", "Mã kho", "Mã vị trí", "Mã lô", "Số lượng", "Giá bán", "Tiền hàng", "Mã nt", "Tỷ giá", "Mã thuế",
           "Tk nợ", "Tk doanh thu", "Tk giá vốn", "Tk thuế có", "Cục thuế", "Vụ việc", "Bộ phận", "Lsx", "Sản phẩm",
           "Hợp đồng", "Phí", "Khế ước", "Nhân viên bán", "Tên KH(thuế)", "Địa chỉ (thuế)", "Mã số Thuế",
           "Nhóm Hàng", "Ghi chú", "Tiền thuế"]

# --- Kiểm tra ngày hết hạn ứng dụng ---
expiration_date = datetime(2025, 6, 26)
current_date = datetime.now()

if current_date > expiration_date:
    st.error("Có lỗi khi chạy chương trình, vui lòng liên hệ tác giả để được hỗ trợ!")
    st.info("Nguyễn Trọng Hoàn - 0902069469")
    st.stop() # Dừng ứng dụng

# --- Hàm trợ giúp chuyển đổi giá trị sang float an toàn ---
def to_float(value):
    """Chuyển đổi giá trị sang float, trả về 0.0 nếu không thể chuyển đổi."""
    try:
        if isinstance(value, str):
            value = value.replace(",", "").strip()
        return float(value)
    except (ValueError, TypeError):
        return 0.0

# --- Hàm làm sạch chuỗi (loại bỏ mọi loại khoảng trắng và chuẩn hóa) ---
def clean_string(s):
    if s is None:
        return ""
    s = re.sub(r'\s+', ' ', str(s)).strip()
    return s

# --- Hàm đọc dữ liệu tĩnh và bảng tra cứu từ Data.xlsx ---
@st.cache_data
def get_static_data_from_excel(file_path):
    """
    Đọc dữ liệu và xây dựng các bảng tra cứu từ Data.xlsx.
    Sử dụng openpyxl để đọc dữ liệu. Kết quả được cache.
    """
    try:
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active

        listbox_data = []
        chxd_detail_map = {}
        store_specific_x_lookup = {}
        
        for row_idx in range(4, ws.max_row + 1):
            row_data_values = [cell.value for cell in ws[row_idx]]

            if len(row_data_values) < 18: continue

            raw_chxd_name = row_data_values[10]
            if raw_chxd_name and clean_string(raw_chxd_name):
                chxd_name_str = clean_string(raw_chxd_name)
                
                if chxd_name_str and chxd_name_str not in listbox_data:
                    listbox_data.append(chxd_name_str)

                g5_val = row_data_values[15] if pd.notna(row_data_values[15]) else None
                f5_val_full = clean_string(row_data_values[16]) if pd.notna(row_data_values[16]) else ''
                h5_val = clean_string(row_data_values[17]).lower() if pd.notna(row_data_values[17]) else ''
                
                if f5_val_full:
                    chxd_detail_map[chxd_name_str] = {
                        'g5_val': g5_val, 'h5_val': h5_val,
                        'f5_val_full': f5_val_full, 'b5_val': chxd_name_str
                    }
                
                # Column mapping for store_specific_x_lookup
                # These indices are based on the Data.xlsx structure for X lookup values
                store_specific_x_lookup[chxd_name_str] = {
                    "xăng e5 ron 92-ii": row_data_values[11], # Original column L
                    "xăng ron 95-iii":   row_data_values[12], # Original column M
                    "dầu do 0,05s-ii":   row_data_values[13], # Original column N
                    "dầu do 0,001s-v":   row_data_values[14]  # Original column O
                }
        
        lookup_table = {} # For "Mã hàng" lookup (I4:J6 in Data.xlsx)
        for row in ws.iter_rows(min_row=4, max_row=7, min_col=9, max_col=10, values_only=True):
            if row[0] and row[1]: lookup_table[clean_string(row[0]).lower()] = row[1]
        
        tmt_lookup_table = {} # For "Thuế bảo vệ môi trường" lookup (I10:J13 in Data.xlsx)
        for row in ws.iter_rows(min_row=10, max_row=13, min_col=9, max_col=10, values_only=True):
            if row[0] and row[1]: tmt_lookup_table[clean_string(row[0]).lower()] = to_float(row[1])
        
        s_lookup_table = {} # For "Tk nợ" lookup (I29:J31 in Data.xlsx)
        for row in ws.iter_rows(min_row=29, max_row=31, min_col=9, max_col=10, values_only=True):
            if row[0] and row[1]: s_lookup_table[clean_string(row[0]).lower()] = row[1]
        
        t_lookup_regular = {} # For "Tk doanh thu" lookup (I33:J35 in Data.xlsx)
        for row in ws.iter_rows(min_row=33, max_row=35, min_col=9, max_col=10, values_only=True):
            if row[0] and row[1]: t_lookup_regular[clean_string(row[0]).lower()] = row[1]
        
        t_lookup_tmt = {} # For "Tk doanh thu" for TMT (I48:J50 in Data.xlsx)
        for row in ws.iter_rows(min_row=48, max_row=50, min_col=9, max_col=10, values_only=True):
            if row[0] and row[1]: t_lookup_tmt[clean_string(row[0]).lower()] = row[1]

        v_lookup_table = {} # For "Tk thuế có" lookup (I53:J55 in Data.xlsx)
        for row in ws.iter_rows(min_row=53, max_row=55, min_col=9, max_col=10, values_only=True):
            if row[0] and row[1]: v_lookup_table[clean_string(row[0]).lower()] = row[1]
        
        u_value = ws['J36'].value # Value from J36 in Data.xlsx
        wb.close()
        
        return {
            "listbox_data": listbox_data, "lookup_table": lookup_table,
            "tmt_lookup_table": tmt_lookup_table, "s_lookup_table": s_lookup_table,
            "t_lookup_regular": t_lookup_regular, "t_lookup_tmt": t_lookup_tmt,
            "v_lookup_table": v_lookup_table, "u_value": u_value,
            "chxd_detail_map": chxd_detail_map, "store_specific_x_lookup": store_specific_x_lookup
        }
    except FileNotFoundError:
        st.error(f"Lỗi: Không tìm thấy file {file_path}. Vui lòng đảm bảo file tồn tại.")
        st.stop()
    except Exception as e:
        st.error(f"Lỗi không mong muốn khi đọc file Data.xlsx: {e}")
        st.exception(e)
        st.stop()

# --- Functions for adding TMT summary row (must be defined before add_summary_row_for_no_invoice) ---
def add_tmt_summary_row(product_name_full, total_bvmt_amount, g5_val, s_lookup, t_lookup_tmt, v_lookup, u_val, h5_val, 
                        representative_date, representative_symbol, total_quantity_for_tmt, tmt_unit_value_for_summary, b5_val, customer_name_for_summary_row, x_lookup_for_store):
    new_tmt_row = [''] * len(headers)
    new_tmt_row[0], new_tmt_row[1], new_tmt_row[2] = g5_val, customer_name_for_summary_row, representative_date
    value_C, value_E = clean_string(representative_date), clean_string(representative_symbol)
    suffix_d = {"Xăng E5 RON 92-II": "1", "Xăng RON 95-III": "2", "Dầu DO 0,05S-II": "3", "Dầu DO 0,001S-V": "4"}.get(product_name_full, "")
    if b5_val == "Nguyễn Huệ": new_tmt_row[3] = f"HNBK{value_C[-2:]}.{value_C[5:7]}.{suffix_d}"
    elif b5_val == "Mai Linh": new_tmt_row[3] = f"MMBK{value_C[-2:]}.{value_C[5:7]}.{suffix_d}"
    else: new_tmt_row[3] = f"{value_E[-2:]}BK{value_C[-2:]}.{value_C[5:7]}.{suffix_d}"
    new_tmt_row[4] = representative_symbol
    new_tmt_row[6], new_tmt_row[7], new_tmt_row[8] = "TMT", "Thuế bảo vệ môi trường", "Lít"
    new_tmt_row[9], new_tmt_row[12] = g5_val, total_quantity_for_tmt
    new_tmt_row[13] = tmt_unit_value_for_summary
    new_tmt_row[14] = round(to_float(total_quantity_for_tmt) * to_float(tmt_unit_value_for_summary), 0) # Tiền hàng for TMT summary
    new_tmt_row[17] = 10
    new_tmt_row[18] = s_lookup.get(h5_val, '')
    new_tmt_row[19] = t_lookup_tmt.get(h5_val, '')
    new_tmt_row[20], new_tmt_row[21] = u_val, v_lookup.get(h5_val, '')
    new_tmt_row[23] = x_lookup_for_store.get(product_name_full.lower(), '')
    new_tmt_row[31] = "" # Tên KH(thuế) cho dòng TMT summary
    new_tmt_row[36] = round(to_float(total_quantity_for_tmt) * to_float(tmt_unit_value_for_summary) * 0.1, 0)
    for idx in [5,10,11,15,16,22,24,25,26,27,28,29,30,32,33,34,35]:
        if idx != 23 and idx < len(new_tmt_row): new_tmt_row[idx] = ''
    return new_tmt_row

# --- Functions for adding summary row for no invoice ---
def add_summary_row_for_no_invoice(data_for_summary_product, bkhd_source_ws, product_name, headers_list,
                    g5_val, b5_val, s_lookup, t_lookup, v_lookup, x_lookup_for_store, u_val, h5_val, common_lookup_table):
    new_row = [''] * len(headers_list)
    new_row[0], new_row[1] = g5_val, f"Khách hàng mua {product_name} không lấy hóa đơn"
    new_row[2] = data_for_summary_product[0][2] if data_for_summary_product else "" # Date (from first row)
    new_row[4] = data_for_summary_product[0][4] if data_for_summary_product else "" # Symbol (from first row)
    value_C, value_E = clean_string(new_row[2]), clean_string(new_row[4])
    suffix_d = {"Xăng E5 RON 92-II": "1", "Xăng RON 95-III": "2", "Dầu DO 0,05S-II": "3", "Dầu DO 0,001S-V": "4"}.get(product_name, "")
    if b5_val == "Nguyễn Huệ": new_row[3] = f"HNBK{value_C[-2:]}.{value_C[5:7]}.{suffix_d}"
    elif b5_val == "Mai Linh": new_row[3] = f"MMBK{value_C[-2:]}.{value_C[5:7]}.{suffix_d}"
    else: new_row[3] = f"{value_E[-2:]}BK{value_C[-2:]}.{value_C[5:7]}.{suffix_d}"
    new_row[5] = f"Xuất bán lẻ theo hóa đơn số {new_row[3]}"
    new_row[6], new_row[7], new_row[8], new_row[9] = common_lookup_table.get(clean_string(product_name).lower(), ''), product_name, "Lít", g5_val
    new_row[10], new_row[11] = '', ''
    total_M = sum(to_float(r[12]) for r in data_for_summary_product) # r[12] is 'Số lượng' from processed row (upsse_row[12])
    new_row[12] = total_M
    new_row[13] = max((to_float(r[13]) for r in data_for_summary_product), default=0.0) # r[13] is 'Giá bán' from processed row (upsse_row[13])

    tien_hang_hd = sum(to_float(r[13]) for r in bkhd_source_ws.iter_rows(min_row=2, values_only=True) if clean_string(r[5]) == "Người mua không lấy hóa đơn" and clean_string(r[8]) == product_name)
    price_per_liter = {"Xăng E5 RON 92-II": 1900, "Xăng RON 95-III": 2000, "Dầu DO 0,05S-II": 1000, "Dầu DO 0,001S-V": 1000}.get(product_name, 0)
    new_row[14] = tien_hang_hd - round(total_M * price_per_liter, 0)

    new_row[15], new_row[16], new_row[17] = '', '', 10
    new_row[18], new_row[19] = s_lookup.get(h5_val, ''), t_lookup.get(h5_val, '')
    new_row[20], new_row[21] = u_val, v_lookup.get(h5_val, '')
    new_row[22] = ''
    new_row[23] = x_lookup_for_store.get(clean_string(product_name).lower(), '')
    for i in range(24, 31): new_row[i] = ''
    new_row[31] = f"Khách mua {product_name} không lấy hóa đơn"
    new_row[32], new_row[33], new_row[34], new_row[35] = "", "", '', ''
    
    TienthueHD_from_original_bkhd = sum(to_float(row_bkhd[14]) for row_bkhd in bkhd_source_ws.iter_rows(min_row=2, values_only=True) if clean_string(row_bkhd[5]) == "Người mua không lấy hóa đơn" and clean_string(row_bkhd[8]) == product_name)
    new_row[36] = TienthueHD_from_original_bkhd - round(total_M * price_per_liter * 0.1, 0) 
    return new_row


def create_per_invoice_tmt_row(original_row_data, tmt_value, g5_val, s_lookup, t_lookup_tmt, v_lookup, u_val, h5_val):
    tmt_row = list(original_row_data)
    tmt_row[6], tmt_row[7], tmt_row[8] = "TMT", "Thuế bảo vệ môi trường", "Lít"
    tmt_row[9] = g5_val
    tmt_row[13] = tmt_value
    tmt_row[14] = round(tmt_value * to_float(original_row_data[12]), 0) # Tiền hàng for TMT row
    tmt_row[17] = 10
    tmt_row[18] = s_lookup.get(h5_val, '')
    tmt_row[19] = t_lookup_tmt.get(h5_val, '')
    tmt_row[20], tmt_row[21] = u_val, v_lookup.get(h5_val, '')
    tmt_row[31] = ""
    tmt_row[36] = round(tmt_value * to_float(original_row_data[12]) * 0.1, 0)
    for idx in [5, 10, 11, 15, 16, 22, 24, 25, 26, 27, 28, 29, 30, 32, 33, 34, 35]:
        if idx < len(tmt_row): tmt_row[idx] = ''
    return tmt_row

# --- Tải dữ liệu tĩnh ---
static_data = get_static_data_from_excel(DATA_FILE_PATH)
listbox_data = static_data["listbox_data"]
lookup_table = static_data["lookup_table"]
tmt_lookup_table = static_data["tmt_lookup_table"]
s_lookup_table = static_data["s_lookup_table"]
t_lookup_regular = static_data["t_lookup_regular"]
t_lookup_tmt = static_data["t_lookup_tmt"]
v_lookup_table = static_data["v_lookup_table"]
u_value = static_data["u_value"]
chxd_detail_map = static_data["chxd_detail_map"]
store_specific_x_lookup = static_data["store_specific_x_lookup"]

# --- Giao diện người dùng Streamlit (ĐÃ CẬP NHẬT) ---
col1, col2 = st.columns([2, 5], vertical_alignment="center")

with col1:
    if os.path.exists(LOGO_PATH):
        st.image(LOGO_PATH, width=180)
with col2:
    st.markdown("""
    <div style="text-align: center;">
        <h2 style="color: red; font-weight: bold; margin: 0; padding: 0; font-size: 26px; line-height: 1.2;">CÔNG TY CỔ PHẦN XĂNG DẦU</h2>
        <h2 style="color: red; font-weight: bold; margin: 0; padding: 0; font-size: 26px; line-height: 1.2;">DẦU KHÍ NAM ĐỊNH</h2>
    </div>
    """, unsafe_allow_html=True)

# Thay thế st.title bằng st.markdown để tùy chỉnh style
st.markdown('<h1 style="text-align: center; color: blue; font-size: 28px;">Công cụ đồng bộ dữ liệu lên phần mềm kế toán SSE</h1>', unsafe_allow_html=True)

st.markdown("""
<style>
@keyframes blinker { 50% { opacity: 0.7; } }
.blinking-warning { padding: 12px; background-color: #FFFACD; border: 1px solid #FFD700; border-radius: 8px; text-align: center; animation: blinker 1.5s linear infinite; }
.blinking-warning p { color: #DC143C; font-weight: bold; margin: 0; font-size: 16px; }
</style>
<div class="blinking-warning">
  <p>Lưu ý quan trọng: Để tránh lỗi, sau khi tải file bảng kê từ POS về, bạn hãy mở lên và lưu lại (ấn Ctrl+S hoặc chọn File/Save) trước khi đưa vào ứng dụng để xử lý.</p>
</div>
<br>
""", unsafe_allow_html=True)

selected_value = st.selectbox("Chọn CHXD:", options=[""] + listbox_data, key='selected_chxd')
uploaded_file = st.file_uploader("Tải lên file bảng kê hóa đơn (.xlsx)", type=["xlsx"])

# --- Footer với thông tin tác giả ---
st.markdown("---") # Thêm đường kẻ ngang để phân tách
st.info("Nếu gặp khó khăn khi sử dụng công cụ, hãy liên hệ Nguyễn Trọng Hoàn - 0902069469")

# --- Xử lý chính ---
if st.button("Xử lý", key='process_button'):
    if not selected_value: st.warning("Vui lòng chọn một giá trị từ danh sách CHXD.")
    elif uploaded_file is None: st.warning("Vui lòng tải lên file bảng kê hóa đơn.")
    else:
        try:
            selected_value_normalized = clean_string(selected_value)
            if selected_value_normalized not in chxd_detail_map:
                st.error(f"Không tìm thấy thông tin chi tiết cho CHXD: '{selected_value_normalized}'")
                st.stop()
            
            chxd_details = chxd_detail_map[selected_value_normalized]
            g5_value, h5_value, f5_value_full, b5_value = chxd_details['g5_val'], chxd_details['h5_val'], chxd_details['f5_val_full'], chxd_details['b5_val']
            x_lookup_for_store = store_specific_x_lookup.get(selected_value_normalized, {})
            if not x_lookup_for_store:
                st.warning(f"Không tìm thấy mã Vụ việc cho cửa hàng '{selected_value_normalized}' trong Data.xlsx.")

            bkhd_wb = load_workbook(uploaded_file)
            bkhd_ws = bkhd_wb.active

            long_cells = [f"H{r_idx+1}" for r_idx, cell in enumerate(bkhd_ws['H']) if cell.value and len(str(cell.value)) > 128]
            if long_cells:
                st.error("Địa chỉ trên ô " + ', '.join(long_cells) + " quá dài, hãy điều chỉnh và thử lại.")
                st.stop()

            all_rows_from_bkhd = list(bkhd_ws.iter_rows(values_only=True))
            temp_bkhd_data = all_rows_from_bkhd[4:] if len(all_rows_from_bkhd) >= 4 else []
            
            vi_tri_cu_idx = [0, 1, 2, 3, 4, 5, 7, 6, 8, 10, 11, 13, 14, 16] 
            
            intermediate_data = []
            for row in temp_bkhd_data:
                if len(row) <= max(vi_tri_cu_idx): continue
                new_row = [row[i] for i in vi_tri_cu_idx]
                if new_row[3]:
                    try: new_row[3] = datetime.strptime(str(new_row[3])[:10], '%d-%m-%Y').strftime('%Y-%m-%d')
                    except ValueError: pass
                ma_kh = new_row[4]
                new_row.append("No" if ma_kh is None or len(clean_string(ma_kh)) > 9 else "Yes")
                intermediate_data.append(new_row)

            if not intermediate_data:
                st.error("Không có dữ liệu hợp lệ trong file bảng kê sau khi xử lý.")
                st.stop()

            b2_bkhd = clean_string(intermediate_data[0][1])
            f5_norm = clean_string(f5_value_full)
            if f5_norm.startswith('1'): f5_norm = f5_norm[1:]
            if f5_norm != b2_bkhd:
                st.error("Bảng kê hóa đơn không phải của cửa hàng bạn chọn.")
                st.stop()

            final_rows, all_tmt_rows = [[''] * len(headers) for _ in range(4)] + [headers], []
            no_invoice_rows = {p: [] for p in ["Xăng E5 RON 92-II", "Xăng RON 95-III", "Dầu DO 0,05S-II", "Dầu DO 0,001S-V"]}

            for row in intermediate_data:
                upsse_row = [''] * len(headers)
                upsse_row[0] = clean_string(row[4]) if row[-1] == 'Yes' and row[4] and clean_string(row[4]) else g5_value
                upsse_row[1], upsse_row[2] = clean_string(row[5]), row[3]
                b_orig, c_orig = clean_string(row[1]), clean_string(row[2])
                if b5_value == "Nguyễn Huệ": upsse_row[3] = f"HN{c_orig[-6:]}"
                elif b5_value == "Mai Linh": upsse_row[3] = f"MM{c_orig[-6:]}"
                else: upsse_row[3] = f"{b_orig[-2:]}{c_orig[-6:]}"
                upsse_row[4] = f"1{b_orig}" if b_orig else ''
                upsse_row[5] = f"Xuất bán lẻ theo hóa đơn số {upsse_row[3]}"
                product_name = clean_string(row[8])
                upsse_row[6], upsse_row[7] = lookup_table.get(product_name.lower(), ''), product_name
                upsse_row[8], upsse_row[9] = "Lít", g5_value
                upsse_row[10], upsse_row[11] = '', ''
                upsse_row[12] = to_float(row[9])
                tmt_value = tmt_lookup_table.get(product_name.lower(), 0.0)
                upsse_row[13] = round(to_float(row[10]) / 1.1 - tmt_value, 2)
                upsse_row[14] = to_float(row[11]) - round(tmt_value * upsse_row[12])
                upsse_row[15], upsse_row[16], upsse_row[17] = '', '', 10
                upsse_row[18] = s_lookup_table.get(h5_value, '')
                upsse_row[19] = t_lookup_regular.get(h5_value, '')
                upsse_row[20], upsse_row[21] = u_value, v_lookup_table.get(h5_value, '')
                upsse_row[22] = ''
                upsse_row[23] = x_lookup_for_store.get(product_name.lower(), '')
                for i in range(24, 31): upsse_row[i] = ''
                upsse_row[31] = upsse_row[1]
                upsse_row[32], upsse_row[33] = row[6], row[7]
                upsse_row[34], upsse_row[35] = '', ''
                upsse_row[36] = to_float(row[12]) - round(upsse_row[12] * tmt_value * 0.1)

                if upsse_row[1] == "Người mua không lấy hóa đơn" and product_name in no_invoice_rows:
                    no_invoice_rows[product_name].append(upsse_row)
                else:
                    final_rows.append(upsse_row)
                    if tmt_value > 0 and upsse_row[12] > 0:
                        all_tmt_rows.append(create_per_invoice_tmt_row(upsse_row, tmt_value, g5_value, s_lookup_table, t_lookup_tmt, v_lookup_table, u_value, h5_value))

            for product_name, rows in no_invoice_rows.items():
                if rows:
                    summary_row = add_summary_row_for_no_invoice(rows, bkhd_ws, product_name, headers, g5_value, b5_value, s_lookup_table, t_lookup_regular, v_lookup_table, x_lookup_for_store, u_value, h5_value, lookup_table)
                    final_rows.append(summary_row)
                    
                    tmt_unit = tmt_lookup_table.get(product_name.lower(), 0)
                    total_qty = sum(to_float(r[12]) for r in rows)
                    customer_name_for_summary_row = summary_row[1]
                    
                    all_tmt_rows.append(add_tmt_summary_row(product_name, 0, g5_value, s_lookup_table, t_lookup_tmt, v_lookup_table, u_value, h5_value, summary_row[2], summary_row[4], total_qty, tmt_unit, b5_value, customer_name_for_summary_row, x_lookup_for_store))

            final_rows.extend(all_tmt_rows)

            up_sse_wb_final = Workbook()
            up_sse_ws_final = up_sse_wb_final.active
            for row_data in final_rows: up_sse_ws_final.append(row_data)

            text_style, date_style = NamedStyle(name="text_style", number_format='@'), NamedStyle(name="date_style", number_format='DD/MM/YYYY')
            exclude_cols = {3, 13, 14, 15, 18, 19, 20, 21, 22, 37}
            
            for r in range(1, up_sse_ws_final.max_row + 1):
                for c in range(1, up_sse_ws_final.max_column + 1):
                    cell = up_sse_ws_final.cell(row=r, column=c)
                    if not cell.value or clean_string(cell.value) == "None": continue
                    if c == 3:
                        try:
                            cell.value = datetime.strptime(clean_string(cell.value), '%Y-%m-%d').date()
                            cell.style = date_style
                        except (ValueError, TypeError):
                            pass
                    elif c not in exclude_cols:
                        cell.style = text_style
            
            for r in range(1, up_sse_ws_final.max_row + 1):
                for c in range(18, 23):
                    up_sse_ws_final.cell(row=r, column=c).number_format = '@'

            up_sse_ws_final.column_dimensions['B'].width = 35
            up_sse_ws_final.column_dimensions['C'].width = 12
            up_sse_ws_final.column_dimensions['D'].width = 12

            output = io.BytesIO()
            up_sse_wb_final.save(output)
            st.success("Đã tạo file UpSSE.xlsx thành công!")
            st.download_button("Tải xuống file UpSSE.xlsx", output.getvalue(), "UpSSE.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        except Exception as e:
            st.error(f"Lỗi trong quá trình xử lý file: {e}")
            st.exception(e)
