import io
import re
from collections import defaultdict
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook, Workbook 
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment # Import thêm các style
# openpyxl.drawing.image và openpyxl.utils.cell không còn cần thiết
# from openpyxl.drawing.image import Image as OpenpyxlImage 
# from openpyxl.utils.cell import coordinate_to_tuple 

# xlsxwriter không còn sử dụng cho hàm này
# import xlsxwriter 

# --- CÁC HÀM TIỆN ÍCH ---
def _clean_string(s):
    """Làm sạch chuỗi, loại bỏ khoảng trắng thừa và ký tự '."""
    if s is None: return ""
    cleaned_s = str(s).strip()
    if cleaned_s.startswith("'"): cleaned_s = cleaned_s[1:]
    return re.sub(r'\s+', ' ', cleaned_s)

def _to_float(value):
    """Chuyển đổi giá trị sang float, xử lý các trường hợp lỗi."""
    if value is None: return 0.0
    try:
        return float(str(value).replace(',', '').strip())
    except (ValueError, TypeError): return 0.0

def _format_number(num):
    """Định dạng số thành chuỗi có dấu phẩy phân cách hàng nghìn và 2 chữ số thập phân.
    Lưu ý: Hàm này chỉ dùng cho các trường hợp không dùng filter Jinja2.
    """
    try:
        return f"{num:,.2f}"
    except (ValueError, TypeError):
        return "0.00"

def _excel_date_to_datetime(excel_date):
    """Chuyển đổi ngày tháng từ định dạng Excel sang đối tượng datetime.
    Hỗ trợ các định dạng số Excel, datetime object, và chuỗi 'dd/mm/yyyy hh:mm:ss', 'dd/mm/yyyy', 'yyyy-mm-dd'.
    """
    if isinstance(excel_date, (int, float)):
        try:
            # openpyxl's data_only=True might convert dates to numbers,
            # pandas can convert Excel numbers to datetime objects
            return pd.to_datetime(excel_date, unit='D', origin='1899-12-30').to_pydatetime()
        except Exception:
            return None
    elif isinstance(excel_date, datetime):
        return excel_date
    elif isinstance(excel_date, str):
        try:
            return datetime.strptime(excel_date, '%d/%m/%Y %H:%M:%S') # Thử định dạng dd/mm/yyyy hh:mm:ss
        except ValueError:
            try:
                return datetime.strptime(excel_date, '%d/%m/%Y') # Thử định dạng dd/mm/yyyy
            except ValueError:
                try:
                    return datetime.strptime(excel_date, '%Y-%m-%d') # Thử định dạng yyyy-mm-dd
                except ValueError:
                    return None
    return None

# --- CÁC HÀM PHÂN TÍCH FILE ---

def _load_discount_data(discount_file_bytes):
    """
    Tải và phân tích dữ liệu chiết khấu từ file Excel 'ChietKhau.xlsx'.
    Trả về một dictionary lồng nhau: {mst_khach_hang: {ten_mat_hang: so_tien_chiet_khau_tren_don_vi}}
    """
    discount_map = defaultdict(dict)
    try:
        wb = load_workbook(io.BytesIO(discount_file_bytes), data_only=True, read_only=True, keep_vba=False, keep_links=False)
        ws = wb.active
        
        # Đọc tên các mặt hàng từ dòng 2, cột D đến G (index 3 đến 6)
        # Ví dụ: Xăng E5 RON 92-II, Xăng RON 95-III, Dầu DO 0.05S-II, Dầu DO 0.001S-V
        product_headers = [_clean_string(cell.value) for cell in ws[2][3:7]]
        
        # Dữ liệu chiết khấu bắt đầu từ dòng 3
        for row_index, row_values in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
            # Đảm bảo có đủ cột cho MST (cột C - index 2) và ít nhất 1 mặt hàng chiết khấu
            if len(row_values) >= 7: # Cột G là index 6, nên cần ít nhất 7 cột (0-6)
                mst_khach_hang = _clean_string(row_values[2]) # Cột C: MST
                
                if mst_khach_hang:
                    # Duyệt qua các cột mặt hàng từ D đến G
                    for i, product_name in enumerate(product_headers):
                        # row_values[3 + i] tương ứng với cột D, E, F, G
                        discount_amount_raw = row_values[3 + i] 
                        
                        if product_name and discount_amount_raw is not None:
                            try:
                                # Mức chiết khấu là số tiền cố định trên mỗi đơn vị
                                discount_amount_per_unit = _to_float(discount_amount_raw)
                                discount_map[mst_khach_hang][product_name] = discount_amount_per_unit
                            except Exception as e:
                                print(f"Cảnh báo: Bỏ qua giá trị chiết khấu lỗi tại dòng {row_index}, cột {chr(68+i)}: {e} - Dữ liệu: '{discount_amount_raw}'")
        wb.close()
    except FileNotFoundError:
        print("Cảnh báo: Không tìm thấy file 'ChietKhau.xlsx'. Chức năng chiết khấu sẽ không hoạt động.")
        return defaultdict(dict) # Trả về map rỗng nếu file không tồn tại
    except Exception as e:
        print(f"Lỗi khi tải file chiết khấu 'ChietKhau.xlsx': {e}")
        # Trả về map rỗng nếu có lỗi để chương trình vẫn chạy
        return defaultdict(dict) 
    return discount_map

def _parse_hddt_file(hddt_bytes):
    """Phân tích dữ liệu từ file Bảng kê HĐĐT."""
    try:
        # Tối ưu hóa: Chỉ đọc dữ liệu, bỏ qua định dạng, VBA, và liên kết
        wb = load_workbook(io.BytesIO(hddt_bytes), data_only=True, read_only=True, keep_vba=False, keep_links=False)
        ws = wb.active
        
        pos_invoices = []
        direct_petroleum_invoices = []
        other_invoices = []

        PETROLEUM_PRODUCTS = [
            'Xăng RON 95-III', 'Dầu DO 0,05S-II', 
            'Xăng E5 RON 92-II', 'Dầu DO 0,001S-V'
        ]
        
        for row_index, row_values in enumerate(ws.iter_rows(min_row=11, values_only=True), start=11):
            quantity = _to_float(row_values[8] if len(row_values) > 8 else None)
            if quantity <= 0:
                continue # Bỏ qua các dòng không có số lượng hoặc số lượng <= 0 (bao gồm dòng tổng)

            fkey = _clean_string(row_values[24] if len(row_values) > 24 else None)
            item_name = _clean_string(row_values[6] if len(row_values) > 6 else None)
            
            # Bổ sung trích xuất Mã số Thuế (cột F - index 5) và Đơn giá (cột J - index 9)
            mst_khach_hang = _clean_string(row_values[5] if len(row_values) > 5 else None)
            unit_price = _to_float(row_values[9] if len(row_values) > 9 else None)
            customer_name = _clean_string(row_values[3] if len(row_values) > 3 else None) # Cột D: Tên khách hàng

            invoice_data = {
                'fkey': fkey,
                'item_name': item_name,
                'quantity': quantity,
                'total_amount': _to_float(row_values[16] if len(row_values) > 16 else None),
                'invoice_number': _clean_string(row_values[19] if len(row_values) > 19 else None),
                'invoice_date_raw': row_values[20] if len(row_values) > 20 else None,
                'invoice_symbol_hddt': _clean_string(row_values[18] if len(row_values) > 18 else None), # Lấy ký hiệu hóa đơn từ cột S (index 18)
                'mst_khach_hang': mst_khach_hang, # Thêm MST khách hàng
                'unit_price': unit_price,         # Thêm đơn giá
                'customer_name': customer_name,   # Thêm Tên khách hàng
                'source_row': row_index
            }

            if fkey and fkey.upper().startswith('POS'):
                pos_invoices.append(invoice_data)
            else:
                if item_name in PETROLEUM_PRODUCTS:
                    direct_petroleum_invoices.append(invoice_data)
                else:
                    other_invoices.append(invoice_data)
        
        wb.close() # Đảm bảo đóng workbook sau khi đọc
        return {
            'pos_invoices': pos_invoices,
            'direct_petroleum_invoices': direct_petroleum_invoices,
            'other_invoices': other_invoices
        }
    except Exception as e:
        raise ValueError(f"Lỗi khi đọc file Bảng kê HĐĐT: {e}")

def _parse_log_bom_file(log_bom_bytes):
    """Phân tích dữ liệu từ file Log Bơm (POS)."""
    try:
        # Tối ưu hóa: Chỉ đọc dữ liệu, bỏ qua định dạng, VBA, và liên kết
        wb = load_workbook(io.BytesIO(log_bom_bytes), data_only=True, read_only=True, keep_vba=False, keep_links=False)
        ws = wb.active
        
        pump_logs = []
        # Các loại giao dịch cần xuất hóa đơn
        VALID_TRANSACTION_TYPES = ['bán lẻ', 'hợp đồng', 'khuyến mãi', 'trả trước']
        
        for row_index, row_values in enumerate(ws.iter_rows(min_row=10, values_only=True), start=10):
            transaction_type = _clean_string(row_values[7] if len(row_values) > 7 else None)
            
            # Kiểm tra xem loại giao dịch có nằm trong danh sách hợp lệ không
            if transaction_type.lower() not in VALID_TRANSACTION_TYPES:
                continue

            fkey = _clean_string(row_values[14] if len(row_values) > 14 else None)
            if not fkey:
                raise ValueError(
                    f"Lỗi nghiêm trọng tại dòng {row_index} của file Log Bơm: "
                    f"Giao dịch '{transaction_type}' bắt buộc phải có mã H.Đơn (FKEY) ở cột O nhưng lại bị trống."
                )

            item_name = _clean_string(row_values[3] if len(row_values) > 3 else None)
            quantity = _to_float(row_values[4] if len(row_values) > 4 else None)
            total_amount = _to_float(row_values[6] if len(row_values) > 6 else None)
            
            # Lấy dữ liệu ngày tháng từ cột B (index 1) của bảng kê POS
            transaction_date_raw = row_values[1] if len(row_values) > 1 else None
            transaction_date_dt = _excel_date_to_datetime(transaction_date_raw)
            # Chuyển đổi về định dạng dd/mm/yyyy
            transaction_date_str = transaction_date_dt.strftime('%d/%m/%Y') if transaction_date_dt else 'N/A'

            pump_logs.append({
                'fkey': fkey,
                'item_name': item_name,
                'quantity': quantity,
                'total_amount': total_amount,
                'source_row': row_index,
                'transaction_date': transaction_date_str # Lưu ngày tháng đã định dạng từ POS
            })
            
        if not pump_logs:
            raise ValueError("Không tìm thấy giao dịch nào cần xuất hóa đơn trong file Log Bơm.")
        
        wb.close() # Đảm bảo đóng workbook sau khi đọc
        return pump_logs
    except Exception as e:
        raise ValueError(f"Lỗi khi đọc file Log Bơm: {e}")


def _generate_discount_report_excel(reconciliation_data, discount_data, template_file_path="BaoCaoChietKhau.xlsx"):
    """
    Điền dữ liệu chênh lệch chiết khấu vào file Excel mẫu và tạo báo cáo.
    Sử dụng openpyxl để ghi dữ liệu vào file mẫu đã có sẵn Table và Slicers.
    Chỉ bao gồm các hóa đơn có discount_match == True.
    """
    try:
        # Load the template workbook
        # data_only=False để giữ lại công thức và các yếu tố khác như bảng, slicer
        output_wb = load_workbook(template_file_path, data_only=False) 
        output_ws = output_wb.active # Get the active sheet from the loaded template

        # Extract selected CHXD name
        selected_chxd_name = reconciliation_data.get('selected_chxd_name', 'N/A')

        # Find min and max dates
        min_date = datetime.max
        max_date = datetime.min
        
        dates_found = False
        for mismatch in reconciliation_data.get('detailed_mismatches', {}).get('amounts', []):
            invoice_date_str = mismatch.get('invoice_date', '') # Already dd/mm/yyyy
            if invoice_date_str and invoice_date_str != 'N/A':
                try:
                    current_date = datetime.strptime(invoice_date_str, '%d/%m/%Y')
                    if current_date < min_date:
                        min_date = current_date
                    if current_date > max_date:
                        max_date = current_date
                    dates_found = True
                except ValueError:
                    # Handle cases where date string might be malformed
                    pass

        # Prepare date range text for A5
        date_range_text = ""
        if dates_found:
            if min_date.date() == max_date.date(): # Same day
                date_range_text = f"Ngày {min_date.day} tháng {min_date.month} năm {min_date.year}"
            else: # Different days
                date_range_text = f"Từ Ngày {min_date.day} tháng {min_date.month} năm {min_date.year} tới Ngày {max_date.day} tháng {max_date.month} năm {max_date.year}"
        else:
            date_range_text = "Không xác định được khoảng thời gian"
        
        # --- ÁP DỤNG ĐỊNH DẠNG FONT VÀ GHI NỘI DUNG VÀO A4, A5 ---
        # Định nghĩa font và cỡ chữ tùy chỉnh cho A4 và A5
        # Bạn có thể điều chỉnh 'Times New Roman' và 12 để khớp với font/cỡ chữ mong muốn
        custom_header_font = Font(name='Times New Roman', size=12, bold=True) 

        # Ghi vào A4
        a4_cell = output_ws.cell(row=4, column=1, value=f"ĐƠN VỊ: CỬA HÀNG XĂNG DẦU {selected_chxd_name.upper()}")
        a4_cell.font = custom_header_font
        
        # Ghi vào A5
        a5_cell = output_ws.cell(row=5, column=1, value=f"Thời gian: {date_range_text}")
        a5_cell.font = custom_header_font

        # Clear existing data from row 11 downwards to prevent old data from remaining
        # We assume that the data starts from row 11 and headers are in row 10.
        if output_ws.max_row > 10:
            # Xóa các hàng từ dòng 11 đến dòng cuối cùng
            output_ws.delete_rows(11, output_ws.max_row - 10) 
        
        # Chuẩn bị dữ liệu để ghi
        report_data_rows = []
        if 'detailed_mismatches' in reconciliation_data and 'amounts' in reconciliation_data['detailed_mismatches']:
            for i, mismatch in enumerate(reconciliation_data['detailed_mismatches']['amounts'], 1):
                if mismatch.get('discount_match') == True:
                    mst_khach_hang = mismatch.get('mst_khach_hang', '')
                    item_name = mismatch.get('item_name', '')
                    don_gia_chiet_khau = discount_data.get(mst_khach_hang, {}).get(item_name, 0.0)

                    report_data_rows.append([
                        i, # STT
                        mismatch.get('customer_name', ''), # Tên khách hàng
                        mst_khach_hang, # MST
                        mismatch.get('invoice_number', ''), # Số hóa đơn
                        mismatch.get('invoice_symbol_hddt', ''), # Ký hiệu hóa đơn
                        mismatch.get('invoice_date', ''), # Ngày tháng
                        item_name, # Mặt hàng
                        mismatch.get('quantity', 0.0), # Số lượng
                        don_gia_chiet_khau, # Đơn giá chiết khấu
                        mismatch.get('actual_difference_amount_raw', 0.0) # Tiền chiết khấu
                    ])

        # Định nghĩa font và cỡ chữ tùy chỉnh cho dữ liệu bảng
        # Bạn có thể điều chỉnh 'Times New Roman' và 10 để khớp với font/cỡ chữ mong muốn
        custom_data_font = Font(name='Times New Roman', size=10) 

        # Ghi dữ liệu vào worksheet, bắt đầu từ dòng 11 (Excel's 1-indexed)
        data_start_row_excel = 11 
        for r_idx, row_data in enumerate(report_data_rows):
            for c_idx, cell_value in enumerate(row_data):
                cell = output_ws.cell(row=data_start_row_excel + r_idx, column=c_idx + 1, value=cell_value)
                
                # Áp dụng font và cỡ chữ cho tất cả các ô dữ liệu
                cell.font = custom_data_font

                # Áp dụng định dạng số cho các cột tương ứng
                # "Số lượng" là cột H (index 7, Excel column 8)
                # "Đơn giá chiết khấu" là cột I (index 8, Excel column 9)
                # "Tiền chiết khấu" là cột J (index 9, Excel column 10)
                if c_idx == 7: # Số lượng
                    cell.number_format = '#,##0.00'
                elif c_idx == 8 or c_idx == 9: # Đơn giá chiết khấu, Tiền chiết khấu
                    cell.number_format = '#,##0'
        
        # Lưu workbook vào buffer
        output_buffer = io.BytesIO()
        output_wb.save(output_buffer)
        output_buffer.seek(0)
        return output_buffer

    except FileNotFoundError:
        raise ValueError(f"Không tìm thấy file mẫu báo cáo chiết khấu: '{template_file_path}'. Vui lòng đảm bảo file tồn tại và có tên đúng.")
    except Exception as e:
        print(f"Lỗi khi tạo báo cáo chiết khấu bằng openpyxl (sử dụng mẫu): {e}")
        raise ValueError(f"Đã xảy ra lỗi khi tạo báo cáo chiết khấu: {e}")

def perform_reconciliation(log_bom_bytes, hddt_bytes, selected_chxd_name, invoice_symbol_from_config, discount_data=None):
    """
    Thực hiện đối soát dữ liệu giữa file Log Bơm (POS) và file Bảng kê HĐĐT.
    Bổ sung bước xác thực CHXD và ký hiệu hóa đơn, và tính toán chiết khấu.
    """
    if discount_data is None:
        discount_data = defaultdict(dict) # Đảm bảo có một dictionary rỗng nếu không có dữ liệu chiết khấu

    try:
        # --- BƯỚC XÁC THỰC CHXD TỪ FILE LOG BƠM (POS) ---
        # Tối ưu hóa: Chỉ đọc dữ liệu, bỏ qua định dạng, VBA, và liên kết
        log_wb = load_workbook(io.BytesIO(log_bom_bytes), data_only=True, read_only=True, keep_vba=False, keep_links=False)
        log_ws = log_wb.active
        
        # Đọc ô A2 (có thể là merged cell A-B-C-D-E2)
        pos_chxd_cell_value = log_ws['A2'].value
        if pos_chxd_cell_value:
            # Loại bỏ "CHXD " và làm sạch chuỗi để so sánh
            pos_chxd_name_extracted = _clean_string(str(pos_chxd_cell_value).replace("CHXD ", ""))
            if pos_chxd_name_extracted.lower() != selected_chxd_name.lower():
                log_wb.close() # Đóng workbook trước khi raise lỗi
                raise ValueError("Bảng kê log bơm không phải của cửa hàng bạn chọn.")
        else:
            log_wb.close() # Đóng workbook trước khi raise lỗi
            raise ValueError("Không tìm thấy thông tin CHXD trong file Log Bơm (ô A2 trống).")
        
        log_wb.close() # Đảm bảo đóng workbook sau khi đọc

        # --- BƯỚC XÁC THỰC KÝ HIỆU HÓA ĐƠN TỪ FILE HĐĐT ---
        # Tối ưu hóa: Chỉ đọc dữ liệu, bỏ qua định dạng, VBA, và liên kết
        hddt_wb = load_workbook(io.BytesIO(hddt_bytes), data_only=True, read_only=True, keep_vba=False, keep_links=False)
        hddt_ws = hddt_wb.active
        
        # Lấy 6 ký tự cuối của ký hiệu hóa đơn từ file cấu hình
        # Đảm bảo ký hiệu từ config đủ dài để cắt
        if len(invoice_symbol_from_config) < 6:
            hddt_wb.close() # Đóng workbook trước khi raise lỗi
            raise ValueError(f"Ký hiệu hóa đơn trong file cấu hình Data_HDDT.xlsx ('{invoice_symbol_from_config}') quá ngắn để xác thực.")
        expected_invoice_symbol_suffix = invoice_symbol_from_config[-6:].upper()

        has_at_least_one_valid_invoice_for_symbol_check = False

        # Duyệt qua cột S (index 18) từ dòng 11 để kiểm tra ký hiệu hóa đơn
        for row_index, row_values in enumerate(hddt_ws.iter_rows(min_row=11, values_only=True), start=11):
            # Kiểm tra xem dòng này có phải là một hóa đơn thực tế (có số lượng > 0) không
            # Cột I (index 8) là số lượng
            quantity_val = _to_float(row_values[8] if len(row_values) > 8 else None)
            
            # Nếu số lượng <= 0, đây có thể là dòng tiêu đề, chân trang, hoặc dòng tổng cộng. Bỏ qua xác thực ký hiệu cho các dòng này.
            if quantity_val <= 0:
                continue

            # Nếu đến đây, đây là một dòng hóa đơn hợp lệ (có số lượng > 0)
            has_at_least_one_valid_invoice_for_symbol_check = True

            # Thực hiện xác thực ký hiệu hóa đơn
            if len(row_values) > 18 and row_values[18] is not None: # Cột S (index 18)
                actual_invoice_symbol_hddt = _clean_string(row_values[18])
                if len(actual_invoice_symbol_hddt) >= 6:
                    if actual_invoice_symbol_hddt[-6:].upper() != expected_invoice_symbol_suffix:
                        hddt_wb.close() # Đóng workbook trước khi raise lỗi
                        raise ValueError("Bảng kê hddt không phải của cửa hàng bạn chọn.")
                else:
                    # Dòng hóa đơn hợp lệ nhưng ký hiệu quá ngắn
                    hddt_wb.close() # Đóng workbook trước khi raise lỗi
                    raise ValueError(f"Ký hiệu hóa đơn tại dòng {row_index} của bảng kê HDDT quá ngắn để xác thực.")
            else:
                # Dòng hóa đơn hợp lệ nhưng thiếu ký hiệu hóa đơn
                hddt_wb.close() # Đóng workbook trước khi raise lỗi
                raise ValueError(f"Hóa đơn tại dòng {row_index} của bảng kê HDDT thiếu ký hiệu hóa đơn (cột S).")
        
        # Sau khi kiểm tra tất cả các dòng, nếu không tìm thấy bất kỳ dòng hóa đơn hợp lệ nào để xác thực ký hiệu.
        if not has_at_least_one_valid_invoice_for_symbol_check:
            hddt_wb.close() # Đóng workbook trước khi raise lỗi
            raise ValueError("Không tìm thấy hóa đơn hợp lệ nào trong file Bảng kê HDDT để xác thực ký hiệu.")
        
        hddt_wb.close() # Đảm bảo đóng workbook sau khi đọc

        # Nếu các bước xác thực thành công, tiếp tục xử lý đối soát
        parsed_hddt_data = _parse_hddt_file(hddt_bytes)
        hddt_invoices = parsed_hddt_data['pos_invoices']
        log_bom_data = _parse_log_bom_file(log_bom_bytes) # log_bom_data giờ đã có 'transaction_date'

        if not hddt_invoices:
             raise ValueError("Không tìm thấy hóa đơn nào có FKEY bắt đầu bằng 'POS' để tiến hành đối soát.")

        log_map = {log['fkey']: log for log in log_bom_data}
        hddt_map = {inv['fkey']: inv for inv in hddt_invoices}

        log_fkeys = set(log_map.keys())
        hddt_fkeys = set(hddt_map.keys())

        missing_invoices_fkeys = sorted(list(log_fkeys - hddt_fkeys))
        extra_invoices_fkeys = sorted(list(hddt_fkeys - log_fkeys))
        common_fkeys = sorted(list(log_fkeys.intersection(hddt_fkeys)))

        quantity_mismatches = []
        amount_mismatches = []

        for fkey in common_fkeys:
            log = log_map[fkey]
            inv = hddt_map[fkey]
            
            # Lấy ngày tháng đã được định dạng từ dữ liệu Log Bơm (POS)
            pos_date_str = log.get('transaction_date', 'N/A')
            
            mismatch_info = {
                'fkey': fkey,
                'invoice_number': inv.get('invoice_number', 'N/A'),
                'invoice_date': pos_date_str, # Sử dụng ngày tháng từ POS
                'hddt_amount': inv['total_amount'], 
                'pos_amount': log['total_amount'],   
                'customer_name': inv.get('customer_name', ''),
                'mst_khach_hang': inv.get('mst_khach_hang', ''),
                'item_name': inv.get('item_name', ''),
                'quantity': inv.get('quantity', 0.0),
                'invoice_symbol_hddt': inv.get('invoice_symbol_hddt', ''),
                'invoice_date_raw': inv['invoice_date_raw'] # Giữ lại để debug nếu cần
            }

            # Kiểm tra chênh lệch số lượng
            if abs(log['quantity'] - inv['quantity']) > 0.001:
                quantity_mismatches.append(mismatch_info)

            # Tính toán chênh lệch thực tế (raw)
            actual_difference_raw = log['total_amount'] - inv['total_amount']
            mismatch_info['actual_difference_amount_raw'] = actual_difference_raw # Lưu giá trị raw để hiển thị và tính toán

            # Kiểm tra chênh lệch thành tiền và tính toán chiết khấu
            # Chênh lệch > 1 VNĐ được coi là có chênh lệch cần kiểm tra
            if abs(actual_difference_raw) > 1: 
                mst_khach_hang = inv.get('mst_khach_hang', '')
                item_name = inv.get('item_name', '')
                quantity = inv.get('quantity', 0.0) # Lấy số lượng từ HDDT

                # Lấy số tiền chiết khấu cố định trên mỗi đơn vị từ dữ liệu chiết khấu
                discount_amount_per_unit = discount_data.get(mst_khach_hang, {}).get(item_name, 0.0)
                
                # Tính toán tổng số tiền chiết khấu dự kiến cho hóa đơn này
                # (Số tiền chiết khấu cố định trên mỗi đơn vị * Số lượng)
                expected_discount_total_amount = round(discount_amount_per_unit * quantity)
                mismatch_info['expected_discount_amount'] = expected_discount_total_amount # Lưu giá trị raw

                # So sánh chênh lệch thực tế với chiết khấu dự kiến
                # Coi là khớp nếu chênh lệch thực tế gần bằng số tiền chiết khấu dự kiến (sai số < 1 VNĐ)
                discount_match = abs(round(actual_difference_raw) - expected_discount_total_amount) < 1 
                mismatch_info['discount_match'] = discount_match
                
                amount_mismatches.append(mismatch_info)
                
        item_summary = defaultdict(lambda: {'quantity': {'pos': 0, 'hddt': 0}, 'amount': {'pos': 0, 'hddt': 0}})
        for log in log_bom_data:
            item_summary[log['item_name']]['quantity']['pos'] += log['quantity']
            item_summary[log['item_name']]['amount']['pos'] += log['total_amount']
        for inv in hddt_invoices:
            item_summary[inv['item_name']]['quantity']['hddt'] += inv['quantity']
            item_summary[inv['item_name']]['amount']['hddt'] += inv['total_amount']
            
        final_item_summary = {}
        for name, data in item_summary.items():
            qty_diff = data['quantity']['pos'] - data['quantity']['hddt']
            amt_diff = data['amount']['pos'] - data['amount']['hddt']
            final_item_summary[name] = {
                'quantity': {'pos': _format_number(data['quantity']['pos']), 'hddt': _format_number(data['quantity']['hddt']), 'difference': _format_number(qty_diff), 'is_match': abs(qty_diff) < 0.001},
                'amount': {'pos': _format_number(data['amount']['pos']), 'hddt': _format_number(data['amount']['hddt']), 'difference': _format_number(amt_diff), 'is_match': abs(amt_diff) < 1}
            }
            
        count_diff = len(log_bom_data) - len(hddt_invoices)
        reconciliation_data = {
            'summary': {
                'pos_count': len(log_bom_data), 
                'hddt_count': len(hddt_invoices), 
                'is_match': count_diff == 0 and not missing_invoices_fkeys and not extra_invoices_fkeys,
                'difference': count_diff,
                'missing_fkeys': missing_invoices_fkeys,
                'extra_fkeys': extra_invoices_fkeys
            },
            'detailed_mismatches': {
                'quantities': quantity_mismatches,
                'amounts': amount_mismatches
            },
            'item_comparison': final_item_summary,
            'non_pos_invoices': {
                'direct_petroleum': parsed_hddt_data['direct_petroleum_invoices'],
                'others': parsed_hddt_data['other_invoices']
            }
        }
        
        return reconciliation_data

    except Exception as e:
        # In ra lỗi chi tiết để debug trên Render logs
        print(f"Lỗi trong quá trình đối soát: {e}")
        raise ValueError(f"Đã xảy ra lỗi trong quá trình đối soát: {e}")

