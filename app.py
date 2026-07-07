import base64
import io
import os
import uuid
import zipfile
import re
import tempfile  # <-- THÊM: Thư viện quản lý thư mục tạm thông minh đa nền tảng
from datetime import datetime
from collections import defaultdict
from flask import Flask, flash, redirect, render_template, request, send_file, url_for, get_flashed_messages, jsonify, session
from openpyxl import load_workbook
import pandas as pd

# --- CÁC IMPORT CHO CÁC HANDLER ---
from detector import detect_report_type
from hddt_handler import process_hddt_report
from pos_handler import process_pos_report
from doisoat_handler import perform_reconciliation, _load_discount_data, _generate_discount_report_excel

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'a_very_strong_and_unified_secret_key')

# --- HÀM TIỆN ÍCH CHO VIỆC NẠP DỮ LIỆU CẤU HÌNH ---
def _clean_string_app(s):
    """Làm sạch chuỗi, loại bỏ khoảng trắng thừa và ký tự '."""
    if s is None:
        return ""
    cleaned_s = str(s).strip()
    if cleaned_s.startswith("'"):
        cleaned_s = cleaned_s[1:]
    return re.sub(r'\s+', ' ', cleaned_s)

def _to_float_app(value):
    """Chuyển đổi giá trị sang float, xử lý các trường hợp lỗi."""
    if value is None:
        return 0.0
    try:
        return float(str(value).replace(',', '').strip())
    except (ValueError, TypeError):
        return 0.0

# --- CUSTOM JINJA2 FILTER ---
@app.template_filter('format_currency')
def format_currency_filter(value):
    """Định dạng số thành chuỗi tiền tệ phân tách hàng nghìn."""
    try:
        num = float(value)
        return f"{num:,.0f}"
    except (ValueError, TypeError):
        return "0"

def _sanitize_filename_piece(s: str) -> str:
    """Làm sạch chuỗi để đưa vào tên file an toàn."""
    if not s:
        return "Untitled"
    s = s.strip()
    return re.sub(r'[\\/:*?"<>|\r\n]+', '_', s)

def _parse_date_like_hddt(cell_val):
    """Phân tích ngày tháng từ bảng kê hóa đơn HDDT."""
    if cell_val is None:
        return None
    if isinstance(cell_val, datetime):
        return cell_val.date()
    if isinstance(cell_val, (int, float)):
        try:
            return pd.to_datetime(float(cell_val), unit='D', origin='1899-12-30').date()
        except Exception:
            return None
    if isinstance(cell_val, str):
        date_str = cell_val.strip()
        fmts = ['%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d', '%Y-%m-%d', '%d/%m/%y', '%d-%m-%y']
        for fmt in fmts:
            try:
                return datetime.strptime(date_str, fmt).date()
            except ValueError:
                continue
    return None

def _parse_date_like_pos(cell_val):
    """Phân tích ngày tháng từ bảng kê hóa đơn POS."""
    if cell_val is None:
        return None
    if isinstance(cell_val, datetime):
        return cell_val.date()
    if isinstance(cell_val, (int, float)):
        try:
            return pd.to_datetime(float(cell_val), unit='D', origin='1899-12-30').date()
        except Exception:
            return None
    if isinstance(cell_val, str):
        date_str = cell_val.strip()
        fmts = [
            '%Y-%m-%d %H:%M:%S', '%Y-%m-%d',
            '%d-%m-%Y %H:%M:%S', '%d-%m-%Y',
            '%d/%m/%Y %H:%M:%S', '%d/%m/%Y'
        ]
        for fmt in fmts:
            try:
                return datetime.strptime(date_str, fmt).date()
            except ValueError:
                continue
    return None

def _extract_report_date_for_filename(file_bytes: bytes, report_type: str, confirmed_date_str: str | None) -> datetime.date:
    """Xác định ngày báo cáo để phục vụ đặt tên tệp tin đầu ra."""
    try:
        if report_type == 'HDDT':
            if confirmed_date_str:
                try:
                    return datetime.strptime(confirmed_date_str, '%Y-%m-%d').date()
                except Exception:
                    pass
            wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
            ws = wb.active
            unique_dates = set()
            for row in ws.iter_rows(min_row=11, values_only=True):
                qty = _to_float_app(row[9] if len(row) > 9 else None)
                if qty > 0:
                    dt = _parse_date_like_hddt(row[21] if len(row) > 21 else None)
                    if dt:
                        unique_dates.add(dt)
            wb.close()
            if unique_dates:
                return min(unique_dates)
        elif report_type == 'POS':
            wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
            ws = wb.active
            unique_dates = set()
            for row in ws.iter_rows(min_row=5, values_only=True):
                dt = _parse_date_like_pos(row[3] if len(row) > 3 else None)
                if dt:
                    unique_dates.add(dt)
            wb.close()
            if unique_dates:
                return min(unique_dates)
    except Exception:
        pass
    return datetime.today().date()

def _make_base_filename(store_name: str, file_date: datetime.date) -> str:
    store = _sanitize_filename_piece(store_name)
    date_part = f"{file_date.day:02d}.{file_date.month:02d}.{file_date.year}"
    return f"{store}.{date_part}"

def load_all_static_config_data():
    """Tải toàn bộ cấu hình từ các tệp Excel tĩnh một lần khi khởi tạo server."""
    static_data = {}
    try:
        wb_hddt = load_workbook("Data_HDDT.xlsx", data_only=True)
        ws_hddt = wb_hddt.active

        chxd_detail_map_pos = {}
        store_specific_x_lookup_pos = {}
        chxd_list_for_hddt = []
        tk_mk_map_hddt = {}
        khhd_map_hddt = {}
        chxd_to_khuvuc_map_hddt = {}
        vu_viec_map_hddt = {}
        chxd_makh_map_hddt = {}

        vu_viec_headers = [_clean_string_app(cell.value) for cell in ws_hddt[2][4:10]]

        for row_idx in range(3, ws_hddt.max_row + 1):
            row_values = [cell.value for cell in ws_hddt[row_idx]]
            if len(row_values) > 11:
                chxd_name = _clean_string_app(row_values[3])
                if chxd_name:
                    chxd_detail_map_pos[chxd_name] = {
                        'g5_val': row_values[10],
                        'h5_val': _clean_string_app(row_values[12]).lower(),
                        'f5_val_full': _clean_string_app(row_values[11]),
                        'b5_val': chxd_name
                    }
                    store_specific_x_lookup_pos[chxd_name] = {
                        "xăng e5 ron 92-ii": row_values[4],
                        "xăng ron 95-iii": row_values[5],
                        "dầu do 0,05s-ii": row_values[6],
                        "dầu do 0,001s-v": row_values[7]
                    }
                    if chxd_name not in chxd_list_for_hddt:
                        chxd_list_for_hddt.append(chxd_name)

                    ma_kho = _clean_string_app(row_values[10])
                    khhd = _clean_string_app(row_values[11])
                    khu_vuc = _clean_string_app(row_values[12])
                    ma_khach_chxd = _clean_string_app(row_values[13]) if len(row_values) > 13 else ''

                    if ma_kho:
                        tk_mk_map_hddt[chxd_name] = ma_kho
                    if khhd:
                        khhd_map_hddt[chxd_name] = khhd
                    if khu_vuc:
                        chxd_to_khuvuc_map_hddt[chxd_name] = khu_vuc
                    if ma_khach_chxd:
                        chxd_makh_map_hddt[chxd_name] = ma_khach_chxd

                    vu_viec_map_hddt[chxd_name] = {}
                    vu_viec_data_row = row_values[4:10]
                    for i, header in enumerate(vu_viec_headers):
                        if header:
                            key = "Dầu mỡ nhờn" if i == len(vu_viec_headers) - 1 else header
                            vu_viec_map_hddt[chxd_name][key] = _clean_string_app(vu_viec_data_row[i])

        def get_lookup_pos_config(min_r, max_r, min_c=1, max_c=2):
            return {
                _clean_string_app(row[0]).lower(): row[1]
                for row in ws_hddt.iter_rows(min_row=min_r, max_row=max_r, min_col=min_c, max_col=min_c + 1, values_only=True)
                if row[0] and row[1] is not None
            }

        tmt_lookup_table_pos = {k: _to_float_app(v) for k, v in get_lookup_pos_config(10, 14).items()}

        static_data['pos_config'] = {
            "lookup_table": get_lookup_pos_config(4, 7),
            "tmt_lookup_table": tmt_lookup_table_pos,
            "s_lookup_table": get_lookup_pos_config(29, 31),
            "t_lookup_regular": get_lookup_pos_config(33, 35),
            "t_lookup_tmt": get_lookup_pos_config(48, 50),
            "v_lookup_table": get_lookup_pos_config(53, 55),
            "u_value": ws_hddt['B36'].value,
            "chxd_detail_map": chxd_detail_map_pos,
            "store_specific_x_lookup": store_specific_x_lookup_pos
        }

        def get_lookup_hddt_config(min_r, max_r, min_c=1, max_c=2):
            return {
                _clean_string_app(row[0]): row[1]
                for row in ws_hddt.iter_rows(min_row=min_r, max_row=max_r, min_col=min_c, max_col=min_c + 1, values_only=True)
                if row[0] and row[1] is not None
            }

        phi_bvmt_map_raw = get_lookup_hddt_config(10, 14)
        phi_bvmt_map_hddt = {_clean_string_app(k): _to_float_app(v) for k, v in phi_bvmt_map_raw.items()}

        static_data['hddt_config'] = {
            "DS_CHXD": chxd_list_for_hddt,
            "tk_mk": tk_mk_map_hddt,
            "khhd_map": khhd_map_hddt,
            "chxd_to_khuvuc_map": chxd_to_khuvuc_map_hddt,
            "vu_viec_map": vu_viec_map_hddt,
            "phi_bvmt_map": phi_bvmt_map_hddt,
            "tk_no_map": get_lookup_hddt_config(29, 31),
            "tk_doanh_thu_map": get_lookup_hddt_config(33, 35),
            "tk_thue_co_map": get_lookup_hddt_config(38, 40),
            "tk_gia_von_value": ws_hddt['B36'].value,
            "tk_no_bvmt_map": get_lookup_hddt_config(44, 46),
            "tk_dt_thue_bvmt_map": get_lookup_hddt_config(48, 50),
            "tk_gia_von_bvmt_value": ws_hddt['B51'].value,
            "tk_thue_co_bvmt_map": get_lookup_hddt_config(53, 55),
            "chxd_makh_map": chxd_makh_map_hddt
        }
        wb_hddt.close()

        wb_mahh = load_workbook("MaHH.xlsx", data_only=True)
        ws_mahh = wb_mahh.active
        ma_hang_map = {}
        petroleum_products_list = []
        for r in ws_mahh.iter_rows(min_row=2, max_col=4, values_only=True):
            ten_hang = _clean_string_app(r[0])
            ma_hang = _clean_string_app(r[2])
            loai_hang = _clean_string_app(r[3])
            if ten_hang and ma_hang:
                ma_hang_map[ten_hang] = ma_hang
            if ten_hang and loai_hang.lower() == 'xăng dầu':
                petroleum_products_list.append(ten_hang)

        static_data['hddt_config']["ma_hang_map"] = ma_hang_map
        static_data['hddt_config']["petroleum_products"] = petroleum_products_list
        static_data['pos_config']["petroleum_products"] = petroleum_products_list
        wb_mahh.close()

        wb_dskh = load_workbook("DSKH.xlsx", data_only=True)
        static_data['hddt_config']["mst_to_makh_map"] = {
            _clean_string_app(r[2]): _clean_string_app(r[3])
            for r in wb_dskh.active.iter_rows(min_row=2, max_col=4, values_only=True)
            if r[2]
        }
        wb_dskh.close()

        try:
            with open("ChietKhau.xlsx", "rb") as f:
                discount_file_bytes = f.read()
            static_data['discount_data'] = _load_discount_data(discount_file_bytes)
        except Exception:
            static_data['discount_data'] = defaultdict(dict)

        return static_data, None
    except Exception as e:
        return None, f"Lỗi nạp dữ liệu cấu hình: {e}"

_global_static_config_data, _static_config_error = load_all_static_config_data()

def get_chxd_list():
    if _static_config_error:
        return []
    chxd_data = []
    for chxd_name, details in _global_static_config_data['pos_config']['chxd_detail_map'].items():
        chxd_data.append({
            'name': chxd_name,
            'symbol': details['f5_val_full']
        })
    chxd_data.sort(key=lambda x: x['name'])
    return chxd_data

@app.route('/', methods=['GET'])
def index():
    """Hiển thị trang chính và khôi phục thông báo, dữ liệu đã lưu từ Session."""
    chxd_list = get_chxd_list()
    active_tab = request.args.get('active_tab', 'upsse')
    
    # Khôi phục thông tin cần xác nhận ngày từ session
    date_ambiguous = session.pop('date_ambiguous', False)
    date_options = session.pop('date_options', [])
    form_data = session.pop('upsse_form_data', {"active_tab": active_tab})
    
    # Khôi phục tệp đang chờ nếu có
    pending_path = session.get('pending_file_path')
    if pending_path and os.path.exists(pending_path):
        try:
            with open(pending_path, 'rb') as f:
                file_bytes = f.read()
            form_data["encoded_file"] = base64.b64encode(file_bytes).decode('utf-8')
        except Exception:
            pass
            
    # Kiểm tra xem có yêu cầu tải xuống file không
    trigger_download = False
    if 'download_file' in session:
        trigger_download = True
        
    return render_template(
        'index.html', 
        chxd_list=chxd_list, 
        form_data=form_data, 
        date_ambiguous=date_ambiguous, 
        date_options=date_options,
        trigger_download=trigger_download
    )

@app.route('/process', methods=['POST'])
def process():
    """Xử lý bảng kê và lưu tệp kết quả tạm thời vào thư mục OS temp, sau đó Redirect về trang chủ."""
    chxd_list = get_chxd_list()
    form_data = {
        "selected_chxd": request.form.get('chxd'),
        "price_periods": request.form.get('price_periods', '1'),
        "invoice_number": request.form.get('invoice_number', '').strip(),
        "confirmed_date": request.form.get('confirmed_date'),
        "encoded_file": request.form.get('file_content_b64')
    }

    # Xóa tệp chờ cũ nếu có để tránh chiếm dụng dung lượng
    old_pending_path = session.pop('pending_file_path', None)
    if old_pending_path and os.path.exists(old_pending_path):
        try:
            os.remove(old_pending_path)
        except Exception:
            pass

    try:
        if _static_config_error:
            raise ValueError(_static_config_error)

        if not form_data["selected_chxd"]:
            flash('Vui lòng chọn CHXD.', 'warning')
            session['upsse_form_data'] = form_data
            return redirect(url_for('index', active_tab='upsse'))

        file_content = None
        if form_data["encoded_file"]:
            file_content = base64.b64decode(form_data["encoded_file"])
        elif 'file' in request.files and request.files['file'].filename != '':
            file_content = request.files['file'].read()
        else:
            flash('Vui lòng tải lên file Bảng kê.', 'warning')
            session['upsse_form_data'] = form_data
            return redirect(url_for('index', active_tab='upsse'))

        # Tự động nhận dạng tệp POS hay HDDT
        report_type = detect_report_type(file_content)
        selected_chxd_symbol = next((x['symbol'] for x in chxd_list if x['name'] == form_data["selected_chxd"]), None)

        if not selected_chxd_symbol:
            raise ValueError(f"Không tìm thấy ký hiệu cho cửa hàng '{form_data['selected_chxd']}'. Vui lòng kiểm tra Data_HDDT.xlsx.")

        if report_type == 'POS':
            result = process_pos_report(
                file_content_bytes=file_content,
                selected_chxd=form_data["selected_chxd"],
                price_periods=form_data["price_periods"],
                new_price_invoice_number=form_data["invoice_number"],
                static_data_pos=_global_static_config_data['pos_config'],
                selected_chxd_symbol=selected_chxd_symbol
            )
        elif report_type == 'HDDT':
            result = process_hddt_report(
                file_content_bytes=file_content,
                selected_chxd=form_data["selected_chxd"],
                price_periods=form_data["price_periods"],
                new_price_invoice_number=form_data["invoice_number"],
                confirmed_date_str=form_data["confirmed_date"],
                static_data_hddt=_global_static_config_data['hddt_config'],
                selected_chxd_symbol=selected_chxd_symbol
            )
        else:
            raise ValueError("Không thể nhận diện tự động loại bảng kê. Vui lòng kiểm tra lại file của bạn.")

        # Nếu cần chọn ngày (đa ngày phát hiện trong tệp) - Sửa đổi đường dẫn lưu tạm thời tương thích đa hệ điều hành
        if isinstance(result, dict) and result.get('choice_needed'):
            temp_pending_path = os.path.join(tempfile.gettempdir(), f"pending_{uuid.uuid4().hex}.dat")
            with open(temp_pending_path, 'wb') as f:
                f.write(file_content)
                
            light_form_data = form_data.copy()
            light_form_data["encoded_file"] = "" # Xóa tệp nặng để tránh lưu quá dung lượng cookie session
            
            session['upsse_form_data'] = light_form_data
            session['date_ambiguous'] = True
            session['date_options'] = result['options']
            session['pending_file_path'] = temp_pending_path
            return redirect(url_for('index', active_tab='upsse'))

        report_date = _extract_report_date_for_filename(file_content, report_type, form_data["confirmed_date"])
        base_filename = _make_base_filename(form_data["selected_chxd"], report_date)

        # Hai giai đoạn giá - Sửa đổi đường dẫn lưu tạm thời tương thích đa hệ điều hành
        if isinstance(result, dict) and ('old' in result or 'new' in result):
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zipf:
                if result.get('old'):
                    result['old'].seek(0)
                    zipf.writestr(f'{base_filename}_GiaCu.xlsx', result['old'].read())
                if result.get('new'):
                    result['new'].seek(0)
                    zipf.writestr(f'{base_filename}_GiaMoi.xlsx', result['new'].read())
            zip_buffer.seek(0)
            
            temp_zip_path = os.path.join(tempfile.gettempdir(), f"processed_{uuid.uuid4().hex}.zip")
            with open(temp_zip_path, 'wb') as f:
                f.write(zip_buffer.read())
                
            session['download_file'] = temp_zip_path
            session['download_name'] = 'UpSSE_2_giai_doan.zip'
            flash('Xử lý Đồng bộ SSE thành công!', 'success')
            return redirect(url_for('index', active_tab='upsse'))

        # Một giai đoạn giá - Sửa đổi đường dẫn lưu tạm thời tương thích đa hệ điều hành
        elif isinstance(result, io.BytesIO):
            temp_xlsx_path = os.path.join(tempfile.gettempdir(), f"processed_{uuid.uuid4().hex}.xlsx")
            with open(temp_xlsx_path, 'wb') as f:
                f.write(result.read())
                
            session['download_file'] = temp_xlsx_path
            session['download_name'] = f'{base_filename}.xlsx'
            flash('Xử lý thành công!', 'success')
            return redirect(url_for('index', active_tab='upsse'))
        else:
            raise ValueError("Hàm xử lý không trả về kết quả hợp lệ.")

    except ValueError as ve:
        flash(str(ve).replace('\n', '<br>'), 'danger')
        session['upsse_form_data'] = form_data
        return redirect(url_for('index', active_tab='upsse'))
    except Exception as e:
        flash(f"Đã xảy ra lỗi không mong muốn: {e}", 'danger')
        session['upsse_form_data'] = form_data
        return redirect(url_for('index', active_tab='upsse'))

@app.route('/download')
def download():
    """Tải tệp tin kết quả về máy và dọn dẹp tệp tin tạm thời trên máy chủ."""
    file_path = session.get('download_file')
    download_name = session.get('download_name', 'export.xlsx')
    if file_path and os.path.exists(file_path):
        try:
            with open(file_path, 'rb') as f:
                file_data = io.BytesIO(f.read())
            os.remove(file_path)
        except Exception:
            file_data = None
            
        session.pop('download_file', None)
        session.pop('download_name', None)
        
        if file_data:
            return send_file(
                file_data, 
                as_attachment=True, 
                download_name=download_name, 
                mimetype='application/octet-stream'
            )
    return redirect(url_for('index'))

@app.route('/reconcile', methods=['POST'])
def reconcile():
    """Xử lý đối soát và chuyển về trang chính để cập nhật kết quả."""
    chxd_list_data = get_chxd_list()
    reconciliation_data = None
    try:
        if _static_config_error:
            raise ValueError(_static_config_error)
        selected_chxd_name = request.form.get('chxd')
        file_log_bom = request.files.get('file_log_bom')
        file_hddt = request.files.get('file_hddt')

        if not selected_chxd_name or not file_log_bom or not file_hddt:
            flash('Vui lòng chọn CHXD và tải đủ 2 tệp tin.', 'warning')
            return redirect(url_for('index', active_tab='doisoat'))

        selected_chxd_symbol = next((x['symbol'] for x in chxd_list_data if x['name'] == selected_chxd_name), None)
        log_bom_bytes = file_log_bom.read()
        hddt_bytes = file_hddt.read()
        discount_data = _global_static_config_data.get('discount_data', defaultdict(dict))

        reconciliation_data = perform_reconciliation(log_bom_bytes, hddt_bytes, selected_chxd_name, selected_chxd_symbol, discount_data)
        if reconciliation_data:
            reconciliation_data['selected_chxd_name'] = selected_chxd_name
            flash('Đối soát thành công!', 'success')
    except Exception as e:
        flash(f"Lỗi trong quá trình đối soát: {e}", 'danger')

    session['upsse_form_data'] = {"active_tab": "doisoat"}
    return render_template('index.html', chxd_list=chxd_list_data, reconciliation_data=reconciliation_data, form_data={"active_tab": "doisoat"}, date_ambiguous=False)

@app.route('/generate_discount_report', methods=['POST'])
def generate_discount_report():
    try:
        reconciliation_data_json = request.json
        discount_data = _global_static_config_data.get('discount_data', defaultdict(dict))
        excel_buffer = _generate_discount_report_excel(reconciliation_data_json, discount_data)
        if excel_buffer:
            excel_buffer.seek(0)
            return send_file(excel_buffer, as_attachment=True, download_name='BaoCaoChietKhau.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/process_stock_card', methods=['POST'])
def process_stock_card():
    chxd_list = get_chxd_list()
    selected_chxd = request.form.get('chxd_thekho')
    try:
        if _static_config_error:
            raise ValueError(_static_config_error)
        uploaded_files = request.files.getlist('files[]')
        excel_buffer = process_stock_card_data(uploaded_files, selected_chxd)
        if excel_buffer:
            excel_buffer.seek(0)
            flash('Xử lý thành công!', 'success')
            return send_file(excel_buffer, as_attachment=True, download_name='TheKho_TuDong.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        flash(f"Lỗi xử lý: {e}", 'danger')
    return redirect(url_for('index', active_tab='thekho'))

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=int(os.environ.get('PORT', 5000)))