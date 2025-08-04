import io
import re
from datetime import datetime
import pandas as pd # Thêm import pandas để xử lý ngày tháng tốt hơn
from openpyxl import load_workbook, Workbook

# --- CÁC HÀM TIỆN ÍCH ---
def _pos_to_float(value):
    """Chuyển đổi giá trị sang float, xử lý các trường hợp lỗi."""
    try:
        if isinstance(value, str):
            value = value.replace(",", "").strip()
        return float(value)
    except (ValueError, TypeError):
        return 0.0

def _pos_clean_string(s):
    """Làm sạch chuỗi, loại bỏ khoảng trắng thừa và ký tự '."""
    if s is None:
        return ""
    return re.sub(r'\s+', ' ', str(s)).strip()

# START: FIX 2 - Cập nhật hàm xử lý ngày tháng để linh hoạt hơn
def _pos_parse_date(date_val):
    """
    Hàm xử lý ngày tháng mạnh mẽ, có khả năng nhận diện nhiều định dạng.
    Trả về đối tượng datetime nếu thành công, None nếu thất bại.
    """
    if date_val is None:
        return None

    if isinstance(date_val, datetime):
        return date_val

    # Xử lý trường hợp ngày tháng là số (chuẩn của Excel)
    if isinstance(date_val, (int, float)):
        try:
            # pandas xử lý số của Excel rất tốt
            return pd.to_datetime(float(date_val), unit='D', origin='1899-12-30').to_pydatetime()
        except (ValueError, TypeError):
            pass

    # Xử lý trường hợp ngày tháng là chuỗi ký tự
    if isinstance(date_val, str):
        date_str = date_val.strip()
        # Các định dạng ngày tháng có thể gặp
        date_formats = [
            '%Y-%m-%d %H:%M:%S', # Định dạng gốc từ POS
            '%Y-%m-%d',         # Chỉ có ngày
            '%d-%m-%Y %H:%M:%S',
            '%d-%m-%Y',
            '%d/%m/%Y %H:%M:%S',
            '%d/%m/%Y',
        ]
        for fmt in date_formats:
            try:
                return datetime.strptime(date_str, fmt)
            except ValueError:
                continue
    
    # Nếu tất cả các cách trên đều thất bại
    return None
# END: FIX 2

# --- HÀM TẠO DÒNG TMT (CHỈ DÙNG CHO HÓA ĐƠN LẺ) ---
def _pos_create_tmt_row_for_individual(original_row, tmt_value, details):
    """Tạo dòng Thuế Bảo vệ Môi trường (TMT) cho hóa đơn riêng lẻ."""
    tmt_row = list(original_row)
    ma_thue_for_calc = _pos_to_float(original_row[17])
    tax_rate_decimal = ma_thue_for_calc / 100.0
    tmt_row[6], tmt_row[7], tmt_row[8] = "TMT", "Thuế bảo vệ môi trường", "Lít"
    tmt_row[9] = details['g5_val']
    tmt_row[13] = tmt_value
    
    tien_hang_bvmt = round(tmt_value * _pos_to_float(original_row[12]))
    tien_thue_bvmt = round(tien_hang_bvmt * tax_rate_decimal)
    tmt_row[14] = tien_hang_bvmt
    tmt_row[36] = tien_thue_bvmt

    tmt_row[18] = details['s_lookup_table'].get(details['h5_val'], '')
    tmt_row[19] = details['t_lookup_tmt'].get(details['h5_val'], '')
    tmt_row[20], tmt_row[21] = details['u_value'], details['v_lookup_table'].get(details['h5_val'], '')
    tmt_row[31] = ""
    for idx in [5, 10, 11, 15, 16, 22, 24, 25, 26, 27, 28, 29, 30, 32, 33, 34, 35]:
        if idx < len(tmt_row): tmt_row[idx] = ''
    return tmt_row

# --- HÀM XỬ LÝ HÓA ĐƠN LẺ ---
def _pos_process_single_row(row, details, selected_chxd):
    """Xử lý một dòng dữ liệu hóa đơn riêng lẻ từ bảng kê POS."""
    upsse_row = [''] * 37
    try:
        ma_kh, ten_kh, ngay_hd_raw, so_ct, so_hd, dia_chi_goc, mst_goc, product_name, so_luong, don_gia_vat, tien_hang_source, tien_thue_source = \
        _pos_clean_string(str(row[4])), _pos_clean_string(str(row[5])), row[3], _pos_clean_string(str(row[1])), _pos_clean_string(str(row[2])), \
        _pos_clean_string(str(row[6])), _pos_clean_string(str(row[7])), _pos_clean_string(str(row[8])), _pos_to_float(row[10]), \
        _pos_to_float(row[11]), _pos_to_float(row[13]), _pos_to_float(row[14])
        ma_thue_percent = _pos_to_float(row[15]) if row[15] is not None else 8.0
    except IndexError:
        raise ValueError("Lỗi đọc cột từ file bảng kê POS. Vui lòng đảm bảo file có đủ các cột từ A đến P.")
    
    upsse_row[0] = ma_kh if ma_kh and len(ma_kh) <= 9 else details['g5_val']
    upsse_row[1] = ten_kh
    
    # START: FIX 2 - Sử dụng hàm xử lý ngày tháng mới
    upsse_row[2] = _pos_parse_date(ngay_hd_raw)
    # END: FIX 2

    if details['b5_val'] == "Nguyễn Huệ": upsse_row[3] = f"HN{so_hd[-6:]}"
    elif details['b5_val'] == "Mai Linh": upsse_row[3] = f"MM{so_hd[-6:]}"
    else: upsse_row[3] = f"{so_ct[-2:]}{so_hd[-6:]}"
    upsse_row[4] = f"1{so_ct}" if so_ct else ''
    upsse_row[5] = f"Xuất bán lẻ theo hóa đơn số {upsse_row[3]}"
    upsse_row[6] = details['lookup_table'].get(product_name.lower(), '')
    upsse_row[7], upsse_row[8] = product_name, "Lít"
    upsse_row[9] = details['g5_val']
    upsse_row[12] = so_luong
    tmt_value = details['tmt_lookup_table'].get(product_name.lower(), 0.0)
    tax_rate_decimal = ma_thue_percent / 100.0
    upsse_row[13] = round(don_gia_vat / (1 + tax_rate_decimal) - tmt_value, 2)
    
    tien_hang_bvmt_le = round(tmt_value * so_luong)
    tien_thue_bvmt_le = round(tien_hang_bvmt_le * tax_rate_decimal)
    upsse_row[14] = tien_hang_source - tien_hang_bvmt_le
    upsse_row[36] = tien_thue_source - tien_thue_bvmt_le
    
    upsse_row[17] = f'{int(ma_thue_percent):02d}'
    upsse_row[18] = details['s_lookup_table'].get(details['h5_val'], '')
    upsse_row[19] = details['t_lookup_regular'].get(details['h5_val'], '')
    upsse_row[20] = details['u_value']
    upsse_row[21] = details['v_lookup_table'].get(details['h5_val'], '')
    upsse_row[23] = details['store_specific_x_lookup'].get(selected_chxd, {}).get(product_name.lower(), '')
    upsse_row[31] = upsse_row[1]
    upsse_row[32] = mst_goc
    upsse_row[33] = dia_chi_goc
    return upsse_row

# --- HÀM TẠO DÒNG TỔNG HỢP ---
def _pos_add_summary_row(original_source_rows, product_name, details, product_tax, selected_chxd, suffix_map):
    """Tạo dòng tổng hợp cho khách vãng lai (người mua không lấy hóa đơn)."""
    total_qty = sum(_pos_to_float(r[10]) for r in original_source_rows)
    total_tien_thue_source = sum(_pos_to_float(r[14]) for r in original_source_rows)
    total_phai_thu = sum(_pos_to_float(r[13]) + _pos_to_float(r[14]) for r in original_source_rows)
    
    tmt_value = details['tmt_lookup_table'].get(product_name.lower(), 0.0)
    tax_rate_decimal = product_tax / 100.0

    tien_hang_dong_bvmt = round(tmt_value * total_qty)
    tien_thue_dong_bvmt = round(tien_hang_dong_bvmt * tax_rate_decimal)
    tien_thue_dong_goc = total_tien_thue_source - tien_thue_dong_bvmt
    tien_hang_dong_goc = total_phai_thu - tien_hang_dong_bvmt - tien_thue_dong_bvmt - tien_thue_dong_goc

    new_row = [''] * 37
    sample_row = original_source_rows[0]
    ngay_hd_raw = sample_row[3]
    so_ct = _pos_clean_string(str(sample_row[1]))
    
    new_row[0] = details['g5_val']
    new_row[1] = f"Khách hàng mua {product_name} không lấy hóa đơn"
    
    # START: FIX 2 - Sử dụng hàm xử lý ngày tháng mới
    new_row[2] = _pos_parse_date(ngay_hd_raw)
    # END: FIX 2

    new_row[4] = f"1{so_ct}" if so_ct else ''
    
    # START: FIX 1 - Cập nhật logic tạo số hóa đơn tổng hợp
    value_E = _pos_clean_string(new_row[4])
    prefix = ""
    if details['b5_val'] == "Nguyễn Huệ": prefix = "HN"
    elif details['b5_val'] == "Mai Linh": prefix = "MM"
    else: prefix = value_E[-2:]

    suffix_d = suffix_map.get(product_name, "")
    date_part = ""
    dt_obj = new_row[2]
    if isinstance(dt_obj, datetime):
        # Tạo định dạng dd.mm theo chuẩn của hddt_handler
        date_part = dt_obj.strftime('%d.%m')
    
    new_row[3] = f"{prefix}BK.{date_part}.{suffix_d}"
    # END: FIX 1
    
    new_row[5] = f"Xuất bán lẻ theo hóa đơn số {new_row[3]}"
    new_row[6] = details['lookup_table'].get(product_name.lower(), '')
    new_row[7], new_row[8] = product_name, "Lít"
    new_row[9] = details['g5_val']
    new_row[12] = total_qty
    
    # START: FIX 3 - Bổ sung tính toán và điền "Giá bán"
    # Lấy đơn giá có VAT từ một dòng mẫu
    don_gia_vat_sample = _pos_to_float(sample_row[11])
    # Áp dụng công thức tính giá bán như hóa đơn lẻ
    new_row[13] = round(don_gia_vat_sample / (1 + tax_rate_decimal) - tmt_value, 2)
    # END: FIX 3

    new_row[14] = tien_hang_dong_goc
    new_row[36] = tien_thue_dong_goc
    
    new_row[17] = f'{int(product_tax):02d}'
    new_row[18] = details['s_lookup_table'].get(details['h5_val'], '')
    new_row[19] = details['t_lookup_regular'].get(details['h5_val'], '')
    new_row[20] = details['u_value']
    new_row[21] = details['v_lookup_table'].get(details['h5_val'], '')
    new_row[23] = details['store_specific_x_lookup'].get(selected_chxd, {}).get(product_name.lower(), '')
    new_row[31] = f"Khách mua {product_name} không lấy hóa đơn"
    
    return new_row, tien_hang_dong_bvmt, tien_thue_dong_bvmt

# --- HÀM TẠO FILE UPPSSE ---
def _pos_generate_upsse_rows(source_data_rows, static_data_pos, selected_chxd, is_new_price_period=False):
    """Tạo các dòng dữ liệu cho file UpSSE từ dữ liệu POS."""
    chxd_details = static_data_pos["chxd_detail_map"].get(selected_chxd)
    if not chxd_details: raise ValueError(f"Không tìm thấy thông tin chi tiết cho CHXD: '{selected_chxd}'")
    details = {**static_data_pos, **chxd_details}
    
    petroleum_products = static_data_pos.get("petroleum_products", [])
    if not petroleum_products:
        print("WARNING: Không tìm thấy mặt hàng nào được đánh dấu là 'Xăng dầu' trong file MaHH.xlsx.")

    no_invoice_rows = {p: [] for p in petroleum_products}

    if is_new_price_period:
        new_price_start_index = len(petroleum_products) + 1
        if new_price_start_index < 5: new_price_start_index = 5
        suffix_map = {product: str(i + new_price_start_index) for i, product in enumerate(petroleum_products)}
    else:
        suffix_map = {product: str(i + 1) for i, product in enumerate(petroleum_products)}

    final_rows, all_tmt_rows = [], []
    product_tax_map = {}
    
    for row_idx, row in enumerate(source_data_rows):
        if not row or row[0] is None: continue
        try:
            ten_kh, product_name, ma_thue_percent = _pos_clean_string(str(row[5])), _pos_clean_string(str(row[8])), _pos_to_float(row[15]) if row[15] is not None else 8.0
        except IndexError: raise ValueError(f"Dòng {row_idx + 5} trong file bảng kê POS không đủ cột.")
        if product_name and product_name not in product_tax_map: product_tax_map[product_name] = ma_thue_percent
        
        if ten_kh == "Người mua không lấy hóa đơn" and product_name in no_invoice_rows:
            no_invoice_rows[product_name].append(row)
        else:
            upsse_row = _pos_process_single_row(row, details, selected_chxd)
            final_rows.append(upsse_row)
            tmt_value = details['tmt_lookup_table'].get(product_name.lower(), 0.0)
            so_luong = _pos_to_float(row[10])
            if tmt_value > 0 and so_luong > 0:
                all_tmt_rows.append(_pos_create_tmt_row_for_individual(upsse_row, tmt_value, details))

    for product, original_rows in no_invoice_rows.items():
        if original_rows:
            product_tax = product_tax_map.get(product, 8.0)
            summary_row, tien_hang_bvmt, tien_thue_bvmt = _pos_add_summary_row(
                original_rows, product, details, product_tax, selected_chxd, suffix_map
            )
            final_rows.append(summary_row)
            
            tmt_unit = details['tmt_lookup_table'].get(product.lower(), 0)
            if tmt_unit > 0 and _pos_to_float(summary_row[12]) > 0:
                tmt_summary = list(summary_row)
                tmt_summary[1] = summary_row[1]
                tmt_summary[6], tmt_summary[7] = "TMT", "Thuế bảo vệ môi trường"
                tmt_summary[13] = tmt_unit
                tmt_summary[18] = details['s_lookup_table'].get(details['h5_val'], '')
                tmt_summary[19] = details['t_lookup_tmt'].get(details['h5_val'], '')
                tmt_summary[20] = details['u_value']
                tmt_summary[21] = details['v_lookup_table'].get(details['h5_val'], '')
                tmt_summary[14] = tien_hang_bvmt
                tmt_summary[36] = tien_thue_bvmt
                for idx in [5, 31, 32, 33]: tmt_summary[idx] = ''
                all_tmt_rows.append(tmt_summary)

    final_rows.extend(all_tmt_rows)
    return final_rows

# --- HÀM TẠO FILE EXCEL ---
def _pos_create_excel_buffer(processed_rows):
    """Tạo một đối tượng BytesIO chứa file Excel từ các dòng dữ liệu đã xử lý."""
    if not processed_rows: return None
    output_wb = Workbook()
    output_ws = output_wb.active
    headers = ["Mã khách", "Tên khách hàng", "Ngày", "Số hóa đơn", "Ký hiệu", "Diễn giải", "Mã hàng", "Tên mặt hàng", "Đvt", "Mã kho", "Mã vị trí", "Mã lô", "Số lượng", "Giá bán", "Tiền hàng", "Mã nt", "Tỷ giá", "Mã thuế", "Tk nợ", "Tk doanh thu", "Tk giá vốn", "Tk thuế có", "Cục thuế", "Vụ việc", "Bộ phận", "Lsx", "Sản phẩm", "Hợp đồng", "Phí", "Khế ước", "Nhân viên bán", "Tên KH(thuế)", "Địa chỉ (thuế)", "Mã số Thuế", "Nhóm Hàng", "Ghi chú", "Tiền thuế"]
    for _ in range(4): output_ws.append([''] * len(headers))
    output_ws.append(headers)
    for r_data in processed_rows: output_ws.append(r_data)
    for row_index in range(6, output_ws.max_row + 1):
        date_cell = output_ws[f'C{row_index}']
        if isinstance(date_cell.value, datetime):
            date_cell.number_format = 'dd/mm/yyyy'
        text_cell = output_ws[f'R{row_index}']
        text_cell.number_format = '@'
    output_ws.column_dimensions['B'].width = 35
    output_ws.column_dimensions['C'].width = 12
    output_ws.column_dimensions['D'].width = 12
    output_buffer = io.BytesIO()
    output_wb.save(output_buffer)
    output_buffer.seek(0)
    return output_buffer

# --- HÀM ĐIỀU PHỐI CHÍNH ---
def process_pos_report(file_content_bytes, selected_chxd, price_periods, new_price_invoice_number, static_data_pos, selected_chxd_symbol, **kwargs):
    """
    Xử lý bảng kê POS để tạo file UpSSE.
    Bao gồm xác thực CHXD dựa trên ký hiệu hóa đơn trong bảng kê POS và mã cửa hàng ở ô B5.
    """
    try:
        if static_data_pos is None:
            raise ValueError("Dữ liệu cấu hình tĩnh cho POS chưa được tải. Vui lòng kiểm tra cấu hình ứng dụng.")

        bkhd_wb = load_workbook(io.BytesIO(file_content_bytes), data_only=True)
        bkhd_ws = bkhd_wb.active
        
        if selected_chxd_symbol is None:
            raise ValueError("Ký hiệu hóa đơn của CHXD chưa được cung cấp để xác thực.")

        if len(selected_chxd_symbol) < 6:
            raise ValueError(f"Ký hiệu hóa đơn trong file cấu hình Data_HDDT.xlsx ('{selected_chxd_symbol}') quá ngắn để xác thực.")
        expected_invoice_symbol_suffix = selected_chxd_symbol[-6:].upper()
        
        found_matching_symbol_in_pos_file = False
        max_rows_to_check = min(bkhd_ws.max_row, 100)
        for row_index, row_values in enumerate(bkhd_ws.iter_rows(min_row=5, max_row=max_rows_to_check, values_only=True), start=5):
            if len(row_values) > 1 and row_values[1] is not None:
                actual_invoice_symbol_pos = _pos_clean_string(row_values[1])
                if len(actual_invoice_symbol_pos) >= 6:
                    if actual_invoice_symbol_pos[-6:].upper() == expected_invoice_symbol_suffix:
                        found_matching_symbol_in_pos_file = True
                        break
        
        if not found_matching_symbol_in_pos_file:
            raise ValueError(f"Bảng kê POS không phải của cửa hàng bạn chọn hoặc không tìm thấy ký hiệu hóa đơn hợp lệ.")

        chxd_details = static_data_pos["chxd_detail_map"].get(selected_chxd)
        if not chxd_details: 
            raise ValueError(f"Không tìm thấy thông tin chi tiết cho CHXD: '{selected_chxd}'. Vui lòng kiểm tra file cấu hình Data_HDDT.xlsx.")
        
        b5_bkhd = _pos_clean_string(str(bkhd_ws['B5'].value))
        f5_norm = _pos_clean_string(chxd_details['f5_val_full'])
        
        if f5_norm and len(f5_norm) >= 6 and f5_norm[-6:] != b5_bkhd:
            raise ValueError(f"Lỗi dữ liệu: Mã cửa hàng không khớp.\n- Mã trong Bảng kê POS (ô B5): '{b5_bkhd}'\n- Mã trong file cấu hình (6 ký tự cuối cột K): '{f5_norm[-6:]}'")
        
        all_source_rows = list(bkhd_ws.iter_rows(min_row=5, values_only=True))
        if price_periods == '1':
            processed_rows = _pos_generate_upsse_rows(all_source_rows, static_data_pos, selected_chxd, is_new_price_period=False)
            if not processed_rows: raise ValueError("Không có dữ liệu hợp lệ để xử lý trong file POS tải lên.")
            return _pos_create_excel_buffer(processed_rows)
        else:
            if not new_price_invoice_number: raise ValueError("Vui lòng nhập 'Số hóa đơn đầu tiên của giá mới' khi chọn 2 giai đoạn giá.")
            split_index = -1
            for i, row in enumerate(all_source_rows):
                if len(row) > 2 and row[2] is not None and _pos_clean_string(str(row[2])) == new_price_invoice_number:
                    split_index = i
                    break
            if split_index == -1: raise ValueError(f"Không tìm thấy số hóa đơn '{new_price_invoice_number}' để chia giai đoạn giá.")
            old_price_rows, new_price_rows = all_source_rows[:split_index], all_source_rows[split_index:]
            buffer_new = _pos_create_excel_buffer(_pos_generate_upsse_rows(new_price_rows, static_data_pos, selected_chxd, is_new_price_period=True))
            buffer_old = _pos_create_excel_buffer(_pos_generate_upsse_rows(old_price_rows, static_data_pos, selected_chxd, is_new_price_period=False))
            if not buffer_new and not buffer_old: raise ValueError("Không có dữ liệu hợp lệ để xử lý trong file POS tải lên.")
            return {'new': buffer_new, 'old': buffer_old}
    except Exception as e:
        raise e
